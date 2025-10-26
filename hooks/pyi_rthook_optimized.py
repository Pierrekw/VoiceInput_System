#!/usr/bin/env python3
"""
优化版PyInstaller运行时钩子 - 轻量版环境检查
"""

import os
import sys
import site
from pathlib import Path

def check_basic_environment():
    """检查基本运行环境"""
    try:
        print("检查基本运行环境...")
    except UnicodeEncodeError:
        print("Checking basic environment...")

    # 检查配置文件
    config_files = ["config.yaml", "voice_correction_dict.txt"]
    for file_name in config_files:
        file_path = Path(file_name)
        if not file_path.exists():
            try:
                print(f"配置文件缺失: {file_name}")
            except UnicodeEncodeError:
                print(f"Missing config file: {file_name}")
        else:
            try:
                print(f"配置文件存在: {file_name}")
            except UnicodeEncodeError:
                print(f"Config file exists: {file_name}")

def check_templates():
    """检查Excel模板文件"""
    template_path = Path("reports/templates")
    if not template_path.exists():
        try:
            print("创建Excel模板目录...")
        except UnicodeEncodeError:
            print("Creating Excel template directory...")
        template_path.mkdir(parents=True, exist_ok=True)

        # 如果模板文件不存在，创建一个基本的模板
        template_file = template_path / "enhanced_measure_template.xlsx"
        if not template_file.exists():
            try:
                print("创建基本Excel模板...")
            except UnicodeEncodeError:
                print("Creating basic Excel template...")
            try:
                import openpyxl
                from openpyxl import Workbook

                wb = Workbook()
                ws = wb.active
                ws.title = "Measurement Report"

                # 基本表头
                headers = ["Standard No", "Standard Content", "Lower Limit", "Upper Limit", "Measurement", "Result", "Deviation", "Voice ID", "Timestamp"]
                for col, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col, value=header)

                # 设置列宽
                for col in range(1, 10):
                    ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15

                wb.save(template_file)
                try:
                    print(f"Excel模板已创建: {template_file}")
                except UnicodeEncodeError:
                    print(f"Excel template created: {template_file}")
            except Exception as e:
                try:
                    print(f"创建Excel模板失败: {e}")
                except UnicodeEncodeError:
                    print(f"Failed to create Excel template: {e}")

def show_model_setup_info():
    """显示模型设置信息"""
    try:
        print("\n" + "=" * 60)
        print("VoiceInput v2.5 Lite - 轻量版")
        print("=" * 60)
        print("轻量版说明:")
        print("- 此版本为轻量版，不包含大型AI模型文件")
        print("- 如需完整语音识别功能，请配置模型文件")
        print("\n模型文件配置:")
        print("1. 将 model/fun 文件夹放置在程序同目录下")
        print("2. 模型文件应包含:")
        print("   - damo/speech_asr_nat-zh-cn_16k-common-vocab8484-pytorch.model")
        print("   - damo/speech_asr_nat-zh-cn_16k-common-vocab8484-pytorch.yaml")
        print("   - damo/fsmn_vad_common-zh-cn-16k-common-pytorch")
        print("   - damo/speech_separation_noh_16k_1684_snapshot.onnx")
        print("   - damo/speech_timestamp_prediction-16k-common-zh-cn-2024-03-14.model.onnx")
        print("\n使用提示:")
        print("- 轻量版适合快速部署和基本功能测试")
        print("- 完整功能需要配置模型文件")
        print("- 程序会自动检测模型文件并相应调整功能")
        print("=" * 60)
    except UnicodeEncodeError:
        # English fallback
        print("\n" + "=" * 60)
        print("VoiceInput v2.5 Lite")
        print("=" * 60)
        print("Lite Version Information:")
        print("- This is a lite version without large AI model files")
        print("- Configure model files for full voice recognition functionality")
        print("\nModel File Configuration:")
        print("1. Place model/fun folder in the same directory as the program")
        print("2. Model files should include:")
        print("   - damo/speech_asr_nat-zh-cn_16k-common-vocab8484-pytorch.model")
        print("   - damo/speech_asr_nat-zh-cn_16k-common-vocab8484-pytorch.yaml")
        print("   - damo/fsmn_vad_common-zh-cn-16k-common-pytorch")
        print("   - damo/speech_separation_noh_16k_1684_snapshot.onnx")
        print("   - damo/speech_timestamp_prediction-16k-common-zh-cn-2024-03-14.model.onnx")
        print("\nUsage Tips:")
        print("- Lite version is suitable for quick deployment and basic testing")
        print("- Full functionality requires model file configuration")
        print("- Program will automatically detect model files and adjust features")
        print("=" * 60)

def main():
    """主函数"""
    try:
        print("初始化轻量版运行环境...")

        # 检查基本环境
        check_basic_environment()

        # 检查模板文件
        check_templates()

        # 显示模型设置信息
        show_model_setup_info()

        # 添加当前目录到Python路径
        current_dir = Path(__file__).parent
        if str(current_dir) not in sys.path:
            sys.path.insert(0, str(current_dir))

        print("轻量版环境初始化完成")
    except UnicodeEncodeError:
        # 如果编码出错，使用英文输出
        print("Initializing lite version environment...")
        current_dir = Path(__file__).parent
        if str(current_dir) not in sys.path:
            sys.path.insert(0, str(current_dir))
        print("Lite version environment initialized.")

if __name__ == "__main__":
    main()