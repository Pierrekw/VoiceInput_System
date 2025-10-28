#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
创建Excel公式示例
演示如何使用Excel公式进行自动判断
"""

import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from datetime import datetime

def create_formula_example():
    """创建包含Excel公式的示例文件"""

    print("🧮 创建Excel公式示例")

    # 创建基础数据
    data = [
        [100, '半径1', 75.5, 85.8, 80.0],  # 在范围内 -> OK
        [200, '半径2', 15.5, 30.5, 25.0],  # 在范围内 -> OK
        [300, '半径3', 8.5, 12.5, 10.0],   # 在范围内 -> OK
        [100, '半径1', 75.5, 85.8, 90.0],  # 超出上限 -> NOK
        [200, '半径2', 15.5, 30.5, 10.0],  # 低于下限 -> NOK
        [400, '半径4', 53.5, 57.5, 55.0],  # 在范围内 -> OK
    ]

    columns = ['标准序号', '标准内容', '下限', '上限', '测量值']

    df = pd.DataFrame(data, columns=columns)

    # 添加空列用于公式计算
    df['判断结果'] = ''
    df['偏差'] = ''
    df['时间戳'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # 保存Excel文件
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"reports/excel_formula_example_{timestamp}.xlsx"
    df.to_excel(filename, index=False)

    # 使用openpyxl添加公式和格式
    workbook = load_workbook(filename)
    worksheet = workbook.active

    # 设置列宽
    column_widths = {
        'A': 10, 'B': 15, 'C': 10, 'D': 10, 'E': 10,
        'F': 12, 'G': 10, 'H': 20
    }
    for col_letter, width in column_widths.items():
        worksheet.column_dimensions[col_letter].width = width

    # 格式化标题行
    for col in range(1, len(columns) + 3):
        cell = worksheet.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    # 添加判断公式和条件格式
    for row in range(2, len(data) + 2):
        # 判断结果公式
        judgment_formula = f'=IF(AND(E{row}>=C{row}, E{row}<=D{row}), "OK", "NOK")'
        worksheet.cell(row=row, column=6, value=judgment_formula)

        # 偏差公式
        deviation_formula = f'=IF(F{row}="OK", MIN(E{row}-C{row}, D{row}-E{row}), IF(E{row}<C{row}, C{row}-E{row}, E{row}-D{row}))'
        worksheet.cell(row=row, column=7, value=deviation_formula)

        # 条件格式：OK显示绿色，NOK显示红色
        judgment_cell = worksheet.cell(row=row, column=6)
        judgment_cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    # 添加条件格式规则
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.styles import Color

    # OK条件：绿色
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    ok_rule = CellIsRule(operator='equal', formula=['"OK"'], fill=green_fill)

    # NOK条件：红色
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    nok_rule = CellIsRule(operator='equal', formula=['"NOK"'], fill=red_fill)

    # 应用条件格式到判断结果列
    worksheet.conditional_formatting.add(f'F2:F{len(data)+1}', ok_rule)
    worksheet.conditional_formatting.add(f'F2:F{len(data)+1}', nok_rule)

    # 保存文件
    workbook.save(filename)
    workbook.close()

    print(f"✅ Excel公式示例已创建: {filename}")
    print("\n📝 包含的公式:")
    print("   判断结果 (F列): =IF(AND(E2>=C2, E2<=D2), \"OK\", \"NOK\")")
    print("   偏差 (G列): =IF(F2=\"OK\", MIN(E2-C2, D2-E2), IF(E2<C2, C2-E2, E2-D2))")
    print("   条件格式: OK显示绿色，NOK显示红色")

    # 验证计算结果
    print("\n📊 公式计算结果:")
    df_result = pd.read_excel(filename)
    for idx, row in df_result.iterrows():
        status = "✅" if row['判断结果'] == "OK" else "❌"
        print(f"   {status} 标准{row['标准序号']}: {row['测量值']} -> {row['判断结果']} (偏差: {row['偏差']})")

    return filename

def create_vlookup_example():
    """创建VLOOKUP查询示例"""

    print("\n🔍 创建VLOOKUP查询示例")

    # 创建测量规范数据表
    spec_data = [
        [100, '半径1', 75.5, 85.8],
        [200, '半径2', 15.5, 30.5],
        [300, '半径3', 8.5, 12.5],
        [400, '半径4', 53.5, 57.5],
        [500, '尺寸1', 10.5, 13.5],
        [600, '尺寸2', 24.25, 28.35],
        [700, '尺寸3', 130.5, 135.5],
    ]

    spec_columns = ['标准序号', '标准内容', '下限', '上限']
    spec_df = pd.DataFrame(spec_data, columns=spec_columns)

    # 创建测量数据表
    measure_data = [
        [100, 80.0],  # 查询标准序号100
        [200, 25.0],  # 查询标准序号200
        [300, 10.0],  # 查询标准序号300
        [100, 90.0],  # 查询标准序号100
        [500, 12.0],  # 查询标准序号500
    ]

    measure_columns = ['标准序号', '测量值']
    measure_df = pd.DataFrame(measure_data, columns=measure_columns)

    # 保存到Excel的不同工作表
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"reports/vlookup_example_{timestamp}.xlsx"

    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        spec_df.to_excel(writer, sheet_name='测量规范', index=False)
        measure_df.to_excel(writer, sheet_name='测量数据', index=False)

    # 使用openpyxl添加VLOOKUP公式
    workbook = load_workbook(filename)

    # 处理测量数据工作表
    worksheet = workbook['测量数据']

    # 添加列标题
    worksheet.cell(row=1, column=3, value='标准内容')
    worksheet.cell(row=1, column=4, value='下限')
    worksheet.cell(row=1, column=5, value='上限')
    worksheet.cell(row=1, column=6, value='判断结果')
    worksheet.cell(row=1, column=7, value='偏差')

    # 添加VLOOKUP公式
    for row in range(2, len(measure_data) + 2):
        standard_id = worksheet.cell(row=row, column=1).value
        measured_value = worksheet.cell(row=row, column=2).value

        # VLOOKUP公式查询测量规范
        content_formula = f'=VLOOKUP(A{row}, 测量规范!$A$2:$D$8, 2, FALSE)'
        lower_formula = f'=VLOOKUP(A{row}, 测量规范!$A$2:$D$8, 3, FALSE)'
        upper_formula = f'=VLOOKUP(A{row}, 测量规范!$A$2:$D$8, 4, FALSE)'

        worksheet.cell(row=row, column=3, value=content_formula)
        worksheet.cell(row=row, column=4, value=lower_formula)
        worksheet.cell(row=row, column=5, value=upper_formula)

        # 判断公式
        judgment_formula = f'=IF(AND(B{row}>=D{row}, B{row}<=E{row}), "OK", "NOK")'
        worksheet.cell(row=row, column=6, value=judgment_formula)

        # 偏差公式
        deviation_formula = f'=IF(F{row}="OK", MIN(B{row}-D{row}, E{row}-B{row}), IF(B{row}<D{row}, D{row}-B{row}, B{row}-E{row}))'
        worksheet.cell(row=row, column=7, value=deviation_formula)

    # 设置列宽
    for worksheet in workbook.worksheets:
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
            worksheet.column_dimensions[col].width = 12

    workbook.save(filename)
    workbook.close()

    print(f"✅ VLOOKUP示例已创建: {filename}")
    print("\n📝 包含的公式:")
    print("   标准内容: =VLOOKUP(A2, 测量规范!$A$2:$D$8, 2, FALSE)")
    print("   下限: =VLOOKUP(A2, 测量规范!$A$2:$D$8, 3, FALSE)")
    print("   上限: =VLOOKUP(A2, 测量规范!$A$2:$D$8, 4, FALSE)")
    print("   判断结果: =IF(AND(B2>=D2, B2<=E2), \"OK\", \"NOK\")")

    return filename

if __name__ == "__main__":
    print("🎯 Excel公式功能演示")
    print("="*60)

    # 创建基础公式示例
    formula_file = create_formula_example()

    # 创建VLOOKUP示例
    vlookup_file = create_vlookup_example()

    print("\n" + "="*60)
    print("📋 功能建议总结:")
    print("="*60)
    print("1. 📊 基础公式方案 - 简单直接，适合固定规范")
    print("2. 🔍 VLOOKUP方案 - 动态查询，适合多规范管理")
    print("3. 🐍 Python集成方案 - 自动化程度高，适合复杂业务")
    print("4. 🎯 混合方案 - Python写入数据，Excel实时计算判断")
    print()
    print("💡 推荐实现步骤:")
    print("1. 使用Python写入基础测量数据")
    print("2. Excel自动查询MeasureSpec获取规范范围")
    print("3. Excel公式自动判断OK/NOK和计算偏差")
    print("4. 条件格式自动着色显示判断结果")