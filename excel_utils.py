#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
增强的Excel导出器 - 支持测量规范和格式化
所有格式化和公式生成都在系统停止时处理，避免识别过程中的性能损失
"""

import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter
import threading
import logging
from typing import Any, List, Tuple, Union, Optional, Dict
from datetime import datetime

# 使用统一的日志工具类
import logging
from utils.logging_utils import LoggingManager

# 获取配置好的日志记录器
logger = LoggingManager.get_logger(
    name='excel_exporter_enhanced',
    level=logging.DEBUG,
    console_level=logging.INFO,
    log_to_console=True,
    log_to_file=True
)

# 新增：导入配置系统
from config_loader import config

class ExcelExporterEnhanced:
    """增强的Excel导出器 - 支持测量规范格式化"""

    def __init__(self, filename: Optional[str] = None, part_no: str = "", batch_no: str = "", inspector: str = ""):
        self._lock: threading.Lock = threading.Lock()

        # 使用配置文件中的文件名，如果未指定则使用默认值
        if filename is None:
            filename = config.get("excel.file_name", "measurement_data.xlsx")
        self.filename: str = filename

        # 保存报告信息
        self.part_no = part_no
        self.batch_no = batch_no
        self.inspector = inspector

        # 内存管理相关属性
        self.voice_id_counter: int = 0
        self.deleted_voice_ids: set = set()
        self.voice_id_to_row: Dict[int, int] = {}
        self.next_insert_row: int = 2
        self.active_record_count: int = 0
        self.current_standard_id: int = 100
        self.template_path: str = config.get("excel.template_path", "reports/enhanced_measure_template.xlsx")

        # 会话数据存储
        self._session_data: List[Tuple[Union[int, str, float], Any, str]] = []

        # 延迟格式化标志
        self._pending_formatting: bool = False

    @staticmethod
    def _float_cell(val: Any) -> float:
        try:
            return float(val)
        except (ValueError, TypeError):
            return 0.0

    @staticmethod
    def _int_cell(val: Any) -> int:
        try:
            return int(ExcelExporterEnhanced._float_cell(val))
        except (ValueError, TypeError):
            return 0

    def create_from_template(self, part_no: str = "", batch_no: str = "", inspector: str = "") -> bool:
        """从模板创建Excel文件"""
        try:
            # 检查模板文件是否存在
            if not os.path.exists(self.template_path):
                logger.warning(f"模板文件不存在: {self.template_path}，使用默认方式")
                self.create_new_file()
                return False

            # 保存报告信息
            self.part_no = part_no
            self.batch_no = batch_no
            self.inspector = inspector

            # 复制模板文件
            import shutil
            shutil.copy2(self.template_path, self.filename)

            # 查找下一个可用的插入位置（跳过模板中的现有数据）
            self._find_next_available_row()

            # 设置标志，表示需要在停止时进行格式化
            self._pending_formatting = True

            logger.info(f"从模板创建Excel文件: {self.filename}")
            logger.info(f"报告信息: 零件号={part_no}, 批次号={batch_no}, 检验员={inspector}")
            logger.info(f"下一个插入位置: 第{self.next_insert_row}行")

            return True

        except Exception as e:
            logger.error(f"从模板创建Excel文件失败: {e}")
            return False

    def _find_next_available_row(self):
        """查找下一个可用的插入位置（跳过模板中的现有数据）"""
        try:
            workbook = load_workbook(self.filename)
            worksheet = workbook.active

            # 从第5行开始查找（因为模板第4行是表头，第5行开始是数据）
            data_start_row = 5

            # 查找第一个完全为空的行
            for row in range(data_start_row, worksheet.max_row + 10):  # 检查到现有数据后面10行
                is_empty = True
                for col in range(1, worksheet.max_column + 1):
                    cell_value = worksheet.cell(row=row, column=col).value
                    if cell_value is not None and str(cell_value).strip():
                        is_empty = False
                        break

                if is_empty:
                    self.next_insert_row = row
                    logger.debug(f"找到空行: {row}")
                    break
            else:
                # 如果没找到空行，使用最大行数+1
                self.next_insert_row = worksheet.max_row + 1
                logger.debug(f"未找到空行，使用最大行数+1: {self.next_insert_row}")

            workbook.close()
        except Exception as e:
            logger.error(f"查找下一个可用行失败: {e}")
            self.next_insert_row = 5  # 回退到默认值

    def create_new_file(self) -> None:
        """创建新Excel文件"""
        try:
            # 创建报告模板数据
            headers = ["测量报告", "", "", "", "", "", "", "", ""]
            info_row = ["零件号: " + self.part_no, "", "批次号: " + self.batch_no, "", "检验员: " + self.inspector, "", "", "", ""]
            data_headers = ["标准序号", "标准内容", "下限", "上限", "测量值", "判断结果", "偏差", "time", "语音录入编号"]

            # 创建DataFrame
            data = [headers, info_row, data_headers]
            df = pd.DataFrame(data)

            # 保存Excel
            df.to_excel(self.filename, index=False, header=False)

            logger.info(f"创建新Excel文件: {self.filename}")

        except Exception as e:
            logger.error(f"创建新Excel文件失败: {e}")

    def append_with_text(
            self,
            data: List[Tuple[Union[float, str], str, str]],  # (数值或文本, 原始语音文本, 处理文本)
            auto_generate_ids: bool = True
        ) -> List[Tuple[int, Union[float, str], str]]:  # 返回 [(ID, 数值或文本, 原始文本)]
            """
            写入带原始语音文本的数据
            延迟格式化处理，提高性能
            """
            if not data:
                logger.warning("没有数据可写入")
                return []

            with self._lock:
                try:
                    # 如果文件不存在，先创建
                    if not os.path.exists(self.filename):
                        self.create_new_file()

                    # 使用openpyxl直接写入数据，避免pandas的格式化开销
                    self._write_data_direct(data)

                    # 标记需要格式化
                    self._pending_formatting = True

                    # 返回写入的记录列表
                    result = []
                    for val, original_text, processed_text in data:
                        if auto_generate_ids:
                            voice_id = self.get_next_voice_id()
                        else:
                            # 对于没有ID的情况，使用-1作为占位符
                            voice_id = -1

                        # 记录到会话数据
                        record_val = val if isinstance(val, str) else self._float_cell(val)
                        self._session_data.append((voice_id, record_val, original_text))

                        result.append((voice_id, record_val, original_text))

                    logger.debug(f"成功写入 {len(result)} 条数据到 {self.filename}")
                    return result

                except Exception as e:
                    logger.error(f"写入Excel失败: {e}")
                    return []

    def _write_data_direct(self, data: List[Tuple[Union[float, str], str, str]]) -> None:
        """直接写入数据，避免格式化开销 - 录音阶段写入record ID + record value + 测量标准序号 + 时间戳"""
        workbook = load_workbook(self.filename)
        worksheet = workbook.active

        # 使用已生成的session_data中的Voice ID，避免重复生成
        # 获取最新的数据，对应本次写入的数据
        start_index = max(0, len(self._session_data) - len(data))
        recent_data = self._session_data[start_index:]

        for i, (voice_id, val, original_text) in enumerate(recent_data):
            # 获取下一个插入位置
            row = self.get_next_insert_position()

            # 录音阶段：写入 record ID + record value + 测量标准序号 + 时间戳
            # 写入标准序号 (第1列)
            worksheet.cell(row=row, column=1, value=self.current_standard_id)

            # 写入语音录入编号 (第10列) - 使用已生成的record ID
            worksheet.cell(row=row, column=10, value=voice_id)

            # 写入测量值 (第6列)
            if isinstance(val, str):
                worksheet.cell(row=row, column=6, value=val)
            else:
                worksheet.cell(row=row, column=6, value=self._float_cell(val))

            # 写入时间戳 (第9列)
            worksheet.cell(row=row, column=9, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))

            # 更新内存映射
            self.voice_id_to_row[voice_id] = row
            self.next_insert_row = row + 1
            self.active_record_count += 1

        workbook.save(self.filename)
        workbook.close()

    def get_next_voice_id(self) -> int:
        """获取下一个语音录入ID"""
        self.voice_id_counter += 1
        # 跳过已删除的ID
        while self.voice_id_counter in self.deleted_voice_ids:
            self.voice_id_counter += 1
        return self.voice_id_counter

    def get_next_insert_position(self) -> int:
        """获取下一个插入位置"""
        return self.next_insert_row

    def finalize_excel_file(self) -> bool:
        """
        最终格式化Excel文件
        在系统停止时调用，添加测量规范查询、判断结果和格式化
        """
        if not self._pending_formatting:
            logger.info("无需格式化Excel文件")
            return True

        try:
            logger.info("🔧 开始最终格式化Excel文件...")

            workbook = load_workbook(self.filename)
            worksheet = workbook.active

            # 1. 更新表头信息
            self._update_header_info(worksheet)

            # 2. 应用测量规范和判断逻辑
            self._apply_measure_spec_logic(worksheet)

            # 3. 应用格式化和样式
            self._apply_formatting_and_styles(worksheet)

            # 4. 添加条件格式
            self._apply_conditional_formatting(worksheet)

            workbook.save(self.filename)
            workbook.close()

            # 清除格式化标志
            self._pending_formatting = False

            logger.info("✅ Excel文件格式化完成")
            return True

        except Exception as e:
            logger.error(f"Excel文件格式化失败: {e}")
            import traceback
            traceback.print_exc()
            return False

    def _update_header_info(self, worksheet: Worksheet) -> None:
        """更新表头信息"""
        try:
            # 不删除模板的第2行，保持模板原样
            # 直接填写报告信息到模板的第2行

            # 检查是否已有第1行（标题行）
            cell_a1 = worksheet.cell(row=1, column=1).value
            if not cell_a1 or str(cell_a1).strip() != "测量报告":
                # 在第1行插入标题
                worksheet.insert_rows(1)
                worksheet.cell(row=1, column=1, value=f"测量报告")

            # 填写报告信息到模板的第2行（保持模板格式不变）
            if worksheet.max_row >= 2:
                # 填写零件号到第2列
                worksheet.cell(row=2, column=2, value=self.part_no)

                # 填写批次号到第4列
                worksheet.cell(row=2, column=4, value=self.batch_no)

                # 填写检验员到第6列
                worksheet.cell(row=2, column=6, value=self.inspector)

                logger.debug(f"填写报告信息: 零件号={self.part_no}, 批次号={self.batch_no}, 检验员={self.inspector}")

            # 不设置任何样式，保持模板原样

        except Exception as e:
            logger.error(f"更新表头信息失败: {e}")

    def _apply_measure_spec_logic(self, worksheet: Worksheet) -> None:
        """应用测量规范查询和判断逻辑"""
        try:
            logger.info(f"🔍 开始应用测量规范逻辑，零件号: {self.part_no}")

            # 查找数据开始行
            data_start_row = self._find_data_start_row(worksheet)
            if data_start_row is None:
                logger.warning("未找到数据开始行")
                return

            logger.debug(f"数据开始行: {data_start_row}, 工作表最大行: {worksheet.max_row}")

            spec_filename = f"{self.part_no}_MeasureSpec.xlsx"
            # 首先在reports/templates目录查找，然后在reports目录查找
            reports_dir = os.path.dirname(self.filename)
            template_dir = os.path.join(reports_dir, "templates")
            spec_path_template = os.path.join(template_dir, spec_filename)
            spec_path_reports = os.path.join(reports_dir, spec_filename)

            logger.debug(f"查找测量规范文件:")
            logger.debug(f"  零件号: {self.part_no}")
            logger.debug(f"  模板目录: {template_dir}")
            logger.debug(f"  报告目录: {reports_dir}")
            logger.debug(f"  模板路径: {spec_path_template}")
            logger.debug(f"  报告路径: {spec_path_reports}")
            logger.debug(f"  模板文件存在: {os.path.exists(spec_path_template)}")
            logger.debug(f"  报告文件存在: {os.path.exists(spec_path_reports)}")

            # 优先使用templates目录中的文件
            if os.path.exists(spec_path_template):
                spec_path = spec_path_template
                logger.debug(f"使用templates目录中的测量规范文件: {spec_path}")
            elif os.path.exists(spec_path_reports):
                spec_path = spec_path_reports
                logger.debug(f"使用reports目录中的测量规范文件: {spec_path}")
            else:
                # 测量规范文件不存在，在Excel中显示警告
                warning_message = f"⚠️ 警告: 未找到零件号 {self.part_no} 的测量规范文件"
                expected_filename = spec_filename
                logger.warning(f"测量规范文件不存在: {spec_path_template} 或 {spec_path_reports}")

                # 在Excel文件中写入警告信息，但不删除现有数据
                # 只在第一行显示警告，保留所有历史识别数据
                worksheet.merge_cells(start_row=data_start_row, start_column=2, end_row=data_start_row, end_column=4)
                warning_cell = worksheet.cell(row=data_start_row, column=2)
                warning_cell.value = f"{warning_message}\n期望文件: {expected_filename}\n(历史识别数据已保留)"
                warning_cell.font = Font(color="FF6600", bold=True, size=10)  # 橙色粗体
                warning_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

                logger.info(f"已在Excel文件中显示测量规范文件缺失警告")
                # 不执行测量规范查询，但继续格式化
                logger.info("跳过测量规范查询，继续执行格式化")
                return  # 正常返回，让Excel文件继续保存

            # 找到了测量规范文件，执行正常的查询逻辑
            # 加载测量规范数据
            logger.info(f"📊 正在加载测量规范数据: {spec_path}")
            spec_data = self._load_measure_spec_data(spec_path)
            if not spec_data:
                logger.warning("测量规范数据加载失败或为空")
                # 在Excel中显示加载失败的警告，但不删除现有数据
                worksheet.merge_cells(start_row=data_start_row, start_column=2, end_row=data_start_row, end_column=4)
                warning_cell = worksheet.cell(row=data_start_row, column=2)
                warning_cell.value = f"⚠️ 警告: 测量规范文件加载失败\n文件: {spec_path}\n请检查文件格式是否正确\n(历史识别数据已保留)"
                warning_cell.font = Font(color="FF6600", bold=True, size=10)  # 橙色粗体
                warning_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                logger.info("已在Excel文件中显示测量规范加载失败警告")
                return  # 正常返回，让Excel文件继续保存

            logger.info(f"✅ 成功加载测量规范数据，包含 {len(spec_data)} 个标准序号")

            # 为每行数据应用规范查询和判断
            updated_count = 0
            for row in range(data_start_row, worksheet.max_row + 1):
                standard_id_cell = worksheet.cell(row=row, column=1)
                if standard_id_cell.value is None:
                    continue

                try:
                    standard_id = int(standard_id_cell.value)
                except (ValueError, TypeError):
                    continue

                # 查找对应的测量规范
                spec_info = self._find_spec_by_id(spec_data, standard_id)
                if spec_info is None:
                    continue

                measured_value_cell = worksheet.cell(row=row, column=6)  # 第6列是测量值
                if measured_value_cell.value is None:
                    continue

                try:
                    measured_value = float(measured_value_cell.value)
                except (ValueError, TypeError):
                    continue

                # 写入序号/Excel ID (第5列) - 在stop阶段填充
                excel_id = row - 4  # 第5行开始，所以excel_id = row - 4
                worksheet.cell(row=row, column=5, value=excel_id)

                # 写入规范信息
                worksheet.cell(row=row, column=2, value=spec_info['content'])  # 标准内容 (第2列)
                worksheet.cell(row=row, column=3, value=spec_info['lower_limit'])  # 下限 (第3列)
                worksheet.cell(row=row, column=4, value=spec_info['upper_limit'])  # 上限 (第4列)

                # 计算判断结果
                judgment = self._calculate_judgment(spec_info, measured_value)
                worksheet.cell(row=row, column=7, value=judgment['result'])  # 判断结果 (第7列)
                worksheet.cell(row=row, column=8, value=judgment['deviation'])  # 偏差 (第8列)

                updated_count += 1

            logger.info(f"应用测量规范逻辑: 更新了{updated_count}行数据")

        except Exception as e:
            logger.error(f"应用测量规范逻辑失败: {e}")
            import traceback
            traceback.print_exc()

    def _load_measure_spec_data(self, spec_path: str) -> Dict[int, Dict[str, Any]]:
        """加载测量规范数据"""
        try:
            workbook = load_workbook(spec_path)
            worksheet = workbook.active

            spec_data = {}
            # 从第2行开始读取数据（跳过标题行）
            for row in range(2, worksheet.max_row + 1):
                standard_id_cell = worksheet.cell(row=row, column=1).value
                if standard_id_cell is None:
                    continue

                try:
                    standard_id = int(standard_id_cell)
                except (ValueError, TypeError):
                    continue

                content = worksheet.cell(row=row, column=2).value or ""
                lower_limit = worksheet.cell(row=row, column=3).value
                upper_limit = worksheet.cell(row=row, column=4).value

                spec_data[standard_id] = {
                    'content': content,
                    'lower_limit': float(lower_limit) if lower_limit is not None else None,
                    'upper_limit': float(upper_limit) if upper_limit is not None else None
                }

            workbook.close()
            return spec_data

        except Exception as e:
            logger.error(f"加载测量规范数据失败: {e}")
            return {}

    def _find_spec_by_id(self, spec_data: Dict[int, Dict[str, Any]], standard_id: int) -> Optional[Dict[str, Any]]:
        """根据标准序号查找测量规范"""
        return spec_data.get(standard_id)

    def _calculate_judgment(self, spec_info: Dict[str, Any], measured_value: float) -> Dict[str, Any]:
        """计算判断结果"""
        lower_limit = spec_info['lower_limit']
        upper_limit = spec_info['upper_limit']

        # 判断结果
        if lower_limit is None and upper_limit is None:
            result = '无规范'
            is_ok = None
            deviation = None
        elif lower_limit is None:  # 只有上限
            if measured_value <= upper_limit:
                result = 'OK'
                is_ok = True
                deviation = upper_limit - measured_value
            else:
                result = 'NOK'
                is_ok = False
                deviation = measured_value - upper_limit
        elif upper_limit is None:  # 只有下限
            if measured_value >= lower_limit:
                result = 'OK'
                is_ok = True
                deviation = measured_value - lower_limit
            else:
                result = 'NOK'
                is_ok = False
                deviation = lower_limit - measured_value
        else:  # 既有上限又有下限
            if lower_limit <= measured_value <= upper_limit:
                result = 'OK'
                is_ok = True
                # 计算到最近的边界的偏差
                deviation_to_lower = measured_value - lower_limit
                deviation_to_upper = upper_limit - measured_value
                deviation = min(deviation_to_lower, deviation_to_upper)
            else:
                result = 'NOK'
                is_ok = False
                # 计算超出范围的偏差
                if measured_value < lower_limit:
                    deviation = lower_limit - measured_value
                else:
                    deviation = measured_value - upper_limit

        return {
            'result': result,
            'is_ok': is_ok,
            'deviation': round(deviation, 2) if deviation is not None else None
        }

    def _apply_formatting_and_styles(self, worksheet: Worksheet) -> None:
        """应用格式化和样式"""
        try:
            # 不修改列宽，保持模板原样

            # 添加边框和全选居中对齐
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )

            center_alignment = Alignment(horizontal="center", vertical="center")

            # 对所有有数据的单元格应用边框和居中对齐
            for row in range(1, worksheet.max_row + 1):
                for col in range(1, worksheet.max_column + 1):
                    cell = worksheet.cell(row=row, column=col)
                    if cell.value is not None:  # 只对有内容的单元格应用格式
                        # 应用边框
                        cell.border = thin_border
                        # 应用居中对齐
                        cell.alignment = center_alignment

            logger.info(f"应用格式化: 边框+居中对齐+列宽调整")

        except Exception as e:
            logger.error(f"应用格式化和样式失败: {e}")

    def _apply_conditional_formatting(self, worksheet: Worksheet) -> None:
        """应用条件格式"""
        try:
            data_start_row = self._find_data_start_row(worksheet)
            if not data_start_row:
                return

            # OK条件：绿色
            green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            ok_rule = CellIsRule(operator='equal', formula=['"OK"'], fill=green_fill)

            # NOK条件：红色
            red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            nok_rule = CellIsRule(operator='equal', formula=['"NOK"'], fill=red_fill)

            # 应用条件格式到判断结果列 (第7列)
            judgment_range = f"G{data_start_row}:G{worksheet.max_row}"
            worksheet.conditional_formatting.add(judgment_range, ok_rule)
            worksheet.conditional_formatting.add(judgment_range, nok_rule)

        except Exception as e:
            logger.error(f"应用条件格式失败: {e}")

    def _find_data_start_row(self, worksheet: Worksheet) -> Optional[int]:
        """查找数据开始行"""
        for row in range(1, worksheet.max_row + 1):
            # 查找包含"标准序号"的行
            for col in range(1, worksheet.max_column + 1):
                cell_value = worksheet.cell(row=row, column=col).value
                if cell_value and '标准序号' in str(cell_value):
                    return row + 1  # 数据从标题行的下一行开始
        return None

    def get_session_data(self) -> List[Tuple[Union[int, str, float], Any, str]]:
        """获取本次会话的所有数据"""
        return self._session_data.copy()

    def clear_session_data(self) -> None:
        """清空会话数据"""
        self._session_data.clear()

# 使用示例
if __name__ == "__main__":
    print("🎯 测试增强Excel导出器")
    print("="*60)

    # 创建导出器实例
    exporter = ExcelExporterEnhanced(
        filename="reports/test_enhanced.xlsx",
        part_no="PART-A001",
        batch_no="B202501",
        inspector="张三"
    )

    # 模拟语音识别数据写入
    test_data = [
        (100, "半径1", "80.0"),
        (200, "半径2", "25.0"),
        (100, "半径1", "90.0"),  # 同一标准序号的重复数据
        (300, "半径3", "10.0")
    ]

    print("📝 模拟写入语音识别数据...")
    results = exporter.append_with_text(test_data)
    print(f"   写入{len(results)}条数据")

    # 模拟系统停止时的最终处理
    print("\n🔧 执行最终格式化...")
    success = exporter.finalize_excel_file()
    print(f"   格式化结果: {'✅ 成功' if success else '❌ 失败'}")

    print(f"\n📁 测试文件: {exporter.filename}")
    print("✅ 所有操作都延迟到停止时处理，避免识别过程中的性能损失")