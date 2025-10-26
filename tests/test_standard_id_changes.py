#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
测试标准序号变更对Excel的影响
验证不同标准序号下的数据写入和显示效果
"""

import sys
import os
import time
import logging
from datetime import datetime

# 添加父目录到路径
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from excel_exporter import ExcelExporter
from logging_utils import LoggingManager

def test_standard_id_changes():
    """测试标准序号变更的完整流程"""

    # 设置日志
    logger = LoggingManager.get_logger(
        name='test_standard',
        level=logging.INFO,
        console_level=logging.INFO,
        log_to_console=True,
        log_to_file=True
    )

    print("="*80)
    print("🧪 测试标准序号变更功能")
    print("="*80)

    # 1. 测试数据准备（按标准序号分组）
    test_scenarios = [
        {
            'standard_id': 100,
            'data': [
                (12.5, "十二点五", "12.5"),
                (15.8, "十五点八", "15.8"),
                (99.2, "九十九点二", "99.2")
            ]
        },
        {
            'standard_id': 200,
            'data': [
                (8.1, "八点一", "8.1"),
                (25.6, "二十五点六", "25.6"),
                (33.3, "三十三点三", "33.3"),
                ("OK", "好的", "OK")
            ]
        },
        {
            'standard_id': 300,
            'data': [
                (101.5, "一百零一点五", "101.5"),
                ("NOK", "不行", "NOK")
            ]
        }
    ]

    part_no = "TEST002"
    batch_no = "B202502"
    inspector = "李四"

    # 2. 创建测试ExcelExporter
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    test_filename = f"reports/test_standard_{timestamp}.xlsx"

    print(f"\n📁 测试文件: {test_filename}")
    print(f"📊 测试场景: {len(test_scenarios)} 个标准序号")

    try:
        # 初始化ExcelExporter
        exporter = ExcelExporter(filename=test_filename)

        # 3. 测试模板复制
        print("\n1️⃣ 测试模板复制功能...")
        success = exporter.create_from_template(part_no, batch_no, inspector)
        if success:
            print("✅ 模板复制成功")
        else:
            print("❌ 模板复制失败，使用默认方式")

        # 4. 按标准序号依次写入数据
        for scenario in test_scenarios:
            standard_id = scenario['standard_id']
            data = scenario['data']

            print(f"\n📋 设置标准序号为 {standard_id}...")
            exporter.current_standard_id = standard_id

            print(f"📝 写入 {len(data)} 条数据...")
            results = exporter.append_with_text(data)

            print(f"   写入结果:")
            for voice_id, value, original in results:
                print(f"   Voice ID: {voice_id}, 值: {value}, 原文: {original}")

        # 5. 测试删除功能（删除中间的数据）
        print(f"\n🗑️ 测试删除功能...")
        # 删除Voice ID 4 (属于标准序号200的数据)
        delete_success = exporter.delete_row_by_voice_id(4)
        if delete_success:
            print("✅ 删除Voice ID 4成功")
        else:
            print("❌ 删除失败")

        # 6. 继续写入新数据到标准序号300
        print(f"\n📝 继续写入标准序号300的新数据...")
        continue_data = [("77.7", "七十七点七", "77.7")]
        continue_results = exporter.append_with_text(continue_data)

        print(f"   继续写入结果:")
        for voice_id, value, original in continue_results:
            print(f"   Voice ID: {voice_id}, 值: {value}, 原文: {original}")

        # 7. 测试重新编号
        print(f"\n🔄 测试Excel重新编号...")
        exporter.renumber_excel_ids()
        print("✅ 重新编号完成")

        # 8. 验证最终结果
        print(f"\n🔍 验证最终结果...")
        verify_standard_id_results(exporter, test_scenarios)

        # 9. 内存状态检查
        print(f"\n🧠 内存状态检查...")
        check_memory_state(exporter)

        print(f"\n✅ 标准序号测试完成！文件保存在: {test_filename}")

    except Exception as e:
        print(f"\n❌ 测试失败: {e}")
        import traceback
        traceback.print_exc()

def verify_standard_id_results(exporter, test_scenarios):
    """验证标准序号变更后的Excel结果"""
    try:
        import openpyxl
        from openpyxl import load_workbook

        # 读取Excel文件
        workbook = load_workbook(exporter.filename)
        worksheet = workbook.active

        print(f"📊 Excel文件内容分析:")

        # 按标准序号分组统计
        standard_groups = {}
        for row in range(4, worksheet.max_row + 1):
            standard_id_cell = worksheet.cell(row=row, column=1)  # A列
            voice_id_cell = worksheet.cell(row=row, column=7)  # G列
            measurement_cell = worksheet.cell(row=row, column=4)  # D列

            if standard_id_cell.value is not None and voice_id_cell.value is not None:
                standard_id = standard_id_cell.value
                voice_id = voice_id_cell.value
                measurement = measurement_cell.value

                if standard_id not in standard_groups:
                    standard_groups[standard_id] = []

                standard_groups[standard_id].append({
                    'row': row,
                    'voice_id': voice_id,
                    'measurement': measurement
                })

        workbook.close()

        # 分析每个标准序号的数据
        for standard_id, records in standard_groups.items():
            print(f"\n📏 标准序号 {standard_id}:")
            print(f"   记录数量: {len(records)}")

            voice_ids = [r['voice_id'] for r in records]
            measurements = [r['measurement'] for r in records]

            print(f"   Voice IDs: {voice_ids}")
            print(f"   测量值: {measurements}")

            # 验证与测试数据的对应关系
            for scenario in test_scenarios:
                if scenario['standard_id'] == standard_id:
                    expected_count = len(scenario['data'])
                    actual_count = len(records)

                    if actual_count == expected_count:
                        print(f"   ✅ 数据量正确: {actual_count}/{expected_count}")
                    else:
                        print(f"   ❌ 数据量错误: {actual_count}/{expected_count}")

        # 验证删除的数据
        deleted_ids = exporter.deleted_voice_ids
        print(f"\n🗑️ 已删除的Voice ID: {deleted_ids}")
        if 4 in deleted_ids:
            print("   ✅ 删除的Voice ID 4在列表中")
        else:
            print("   ❌ 删除的Voice ID 4不在列表中")

        # 检查删除的数据是否真的不在文件中
        all_voice_ids = set()
        for records in standard_groups.values():
            for r in records:
                all_voice_ids.add(r['voice_id'])

        missing_deleted = deleted_ids - all_voice_ids
        if missing_deleted == deleted_ids:
            print("   ✅ 删除的ID已正确从文件中移除")
        else:
            print(f"   ❌ 删除的ID仍有部分在文件中: {missing_deleted}")

    except Exception as e:
        print(f"❌ 验证失败: {e}")

def check_memory_state(exporter):
    """检查内存状态"""
    print(f"🧠 内存状态:")
    print(f"   voice_id_counter: {exporter.voice_id_counter}")
    print(f"   next_insert_row: {exporter.next_insert_row}")
    print(f"   active_record_count: {exporter.active_record_count}")
    print(f"   current_standard_id: {exporter.current_standard_id}")
    print(f"   deleted_voice_ids: {exporter.deleted_voice_ids}")

    # 验证映射一致性
    total_records = len(exporter.voice_id_to_row) + len(exporter.deleted_voice_ids)
    if total_records == exporter.voice_id_counter:
        print("✅ 内存映射一致性正确")
    else:
        print(f"❌ 内存映射不一致: 总记录数{total_records} != 计数器{exporter.voice_id_counter}")

    # 显示当前标准序号
    print(f"🎯 当前标准序号: {exporter.current_standard_id}")

if __name__ == "__main__":
    test_standard_id_changes()