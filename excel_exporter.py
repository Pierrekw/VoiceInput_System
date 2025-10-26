# -*- coding: utf-8 -*-
# Excel Exporter Module - Optimized Version
from __future__ import annotations
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, Alignment
import threading
import logging
from typing import Any, List, Tuple, Union, Optional, Dict, Set

# 使用统一的日志工具类
import logging
from logging_utils import LoggingManager

# 获取配置好的日志记录器
logger = LoggingManager.get_logger(
    name='excel_exporter',
    level=logging.DEBUG,  # 文件记录DEBUG级别
    console_level=logging.INFO,  # 控制台显示INFO级别
    log_to_console=True,
    log_to_file=True
)

# 新增：导入配置系统
from config_loader import config

class ExcelExporter:
    def __init__(self, filename: Optional[str] = None):
        self._lock: threading.Lock = threading.Lock()
        # 使用配置文件中的文件名，如果未指定则使用默认值
        if filename is None:
            filename = config.get("excel.file_name", "measurement_data.xlsx")
        self.filename: str = filename
        # 使用配置系统获取是否包含原始语音的设置
        include_original = config.get("excel.formatting.include_original", True)

        # 根据header_language配置设置列名
        header_language = config.get("excel.formatting.header_language", "zh")
        self.columns = []

        if header_language == "en":
            self.columns.extend(["Standard ID", "Excel ID", "Measurement", "Timestamp", "Processed Text", "Voice ID"])
        else:
            self.columns.extend(["标准序号", "Excel编号", "测量值", "时间戳", "处理文本", "语音录入编号"])

        # 根据配置决定是否添加原始语音列
        if include_original:
            if header_language == "en":
                self.columns.append("Original Text")
            else:
                self.columns.append("原始语音")

        # ⭐ 新增：内存行号管理属性
        self.voice_id_counter: int = 0  # 语音录入ID计数器
        self.deleted_voice_ids: Set[int] = set()  # 已删除的voice_id集合
        self.voice_id_to_row: Dict[int, int] = {}  # voice_id -> excel_row 映射
        self.next_insert_row: int = 4  # 下一个插入行号（从第4行开始，前3行是表头）
        self.active_record_count: int = 0  # 当前有效记录数量
        self.current_standard_id: int = 100  # 当前标准序号
        self.template_path: str = "reports/report_template.xlsx"  # 模板文件路径

        self._last_id: int = 0  # Excel编号计数器（停止时重新编号）
        self._initialize_last_id()
        # 记录本次会话的所有数据
        self._session_data: List[Tuple[Union[int, str, float], Any, str]] = []

    @staticmethod
    def _float_cell(val: Any) -> float:
        try:
            return float(val)
        except (ValueError, TypeError):
            return 0.0
    
    @staticmethod
    def _int_cell(val: Any) -> int:
        try:
            return int(ExcelExporter._float_cell(val))   # 先转 float 再 int，容忍 3.0
        except (ValueError, TypeError):
            return 0
       
    
    def _initialize_last_id(self) -> None:
        if os.path.exists(self.filename):
            try:
                workbook = load_workbook(self.filename, read_only=True)
                worksheet = workbook.active
                if worksheet is None:         
                    return                    
                max_row = worksheet.max_row
                if max_row > 1:
                    id_values: List[int] = []
                    for row in range(2, max_row + 1):
                        cell_value = worksheet.cell(row=row, column=1).value
                        if cell_value is not None:
                            try:
                                id_values.append(self._int_cell(cell_value))
                            except (ValueError, TypeError):
                                continue
                    self._last_id = max(id_values) if id_values else 0
                else:
                    self._last_id = 0
                workbook.close()
            except Exception as e:
                logger.error(f"使用openpyxl读取Excel文件时出错: {e}")
                try:
                    df = pd.read_excel(self.filename)
                    if not df.empty and "编号" in df.columns:
                        self._last_id = self._int_cell(df["编号"].max())
                    else:
                        self._last_id = 0
                except Exception as fallback_error:
                    logger.error(f"回退到pandas也失败: {fallback_error}")
                    self._last_id = 0

    def get_next_id(self) -> int:
        # ❌ 删除此方法，改为停止时重新编号
        # self._last_id += 1
        # return self._last_id
        logger.warning("get_next_id()已弃用，请使用renumber_excel_ids()进行重新编号")
        return self._last_id

    # ⭐ 新增方法：获取下一个语音录入ID
    def get_next_voice_id(self) -> int:
        """
        获取下一个语音录入ID（永远递增）
        用于G列的语音录入编号
        """
        self.voice_id_counter += 1
        return self.voice_id_counter

    # ⭐ 新增方法：获取下一个插入位置
    def get_next_insert_position(self) -> Tuple[int, int]:
        """
        获取下一个数据插入位置（纯内存操作，性能极佳）
        返回: (voice_id, excel_row)
        """
        voice_id = self.get_next_voice_id()
        excel_row = self.next_insert_row

        # 更新内存状态
        self.voice_id_to_row[voice_id] = excel_row
        self.next_insert_row += 1
        self.active_record_count += 1

        logger.debug(f"分配插入位置: voice_id={voice_id}, excel_row={excel_row}")
        return voice_id, excel_row

    # ⭐ 新增方法：从模板创建文件
    def create_from_template(self, part_no: str, batch_no: str, inspector: str) -> bool:
        """
        从模板创建新的Excel文件

        Args:
            part_no: 零件号
            batch_no: 批次号
            inspector: 检验员姓名

        Returns:
            bool: 创建成功返回True
        """
        try:
            if not os.path.exists(self.template_path):
                logger.warning(f"模板文件不存在: {self.template_path}，使用默认创建方式")
                self.create_new_file()
                return False

            # 复制模板文件
            import shutil
            shutil.copy2(self.template_path, self.filename)
            logger.info(f"从模板创建Excel文件: {self.filename}")

            # 写入表头信息
            self.write_header_info(part_no, batch_no, inspector)

            # 初始化内存状态
            self.next_insert_row = 4  # 从第4行开始写入数据
            self.active_record_count = 0

            return True

        except Exception as e:
            logger.error(f"从模板创建Excel文件失败: {e}")
            # 降级到默认创建方式
            self.create_new_file()
            return False

    # ⭐ 新增方法：写入表头信息
    def write_header_info(self, part_no: str, batch_no: str, inspector: str) -> None:
        """
        写入表头信息（零件号、批次号、检验员）

        Args:
            part_no: 零件号
            batch_no: 批次号
            inspector: 检验员姓名
        """
        try:
            import openpyxl
            from openpyxl import load_workbook

            workbook = load_workbook(self.filename)
            worksheet = workbook.active

            # 写入B1: 零件号
            worksheet.cell(row=1, column=2, value=part_no)

            # 写入D1: 批次号
            worksheet.cell(row=1, column=4, value=batch_no)

            # 写入F1: 检验员
            worksheet.cell(row=1, column=6, value=inspector)

            workbook.save(self.filename)
            logger.info(f"表头信息已写入: 零件号={part_no}, 批次号={batch_no}, 检验员={inspector}")

        except Exception as e:
            logger.error(f"写入表头信息失败: {e}")

    # ⭐ 新增方法：删除指定语音ID的行
    def delete_row_by_voice_id(self, voice_id: int) -> bool:
        """
        删除指定语音ID的行

        Args:
            voice_id: 要删除的语音ID

        Returns:
            bool: 删除成功返回True
        """
        try:
            # 检查是否存在
            if voice_id not in self.voice_id_to_row:
                logger.warning(f"voice_id {voice_id} 不存在")
                return False

            excel_row = self.voice_id_to_row[voice_id]

            # 删除Excel行
            self._delete_excel_row_by_number(excel_row)

            # 更新内存状态
            self.deleted_voice_ids.add(voice_id)
            del self.voice_id_to_row[voice_id]
            self.active_record_count -= 1

            # 重新计算行号映射
            self._recalculate_row_mappings_after_deletion(excel_row)

            # 更新下一个插入行号
            self.next_insert_row = max(self.voice_id_to_row.values(), default=3) + 1

            logger.info(f"删除完成: voice_id={voice_id}, 下一插入行={self.next_insert_row}")
            return True

        except Exception as e:
            logger.error(f"删除voice_id {voice_id}失败: {e}")
            return False

    # ⭐ 新增方法：删除后重新计算行号映射
    def _recalculate_row_mappings_after_deletion(self, deleted_row: int) -> None:
        """
        删除后重新计算行号映射
        删除第5行后，原来第6行变成第5行，第7行变成第6行...
        """
        updated_mappings = {}

        for voice_id, old_row in self.voice_id_to_row.items():
            if old_row > deleted_row:
                # 行号在删除行之后的，需要减1
                updated_mappings[voice_id] = old_row - 1
            else:
                # 行号在删除行之前的，保持不变
                updated_mappings[voice_id] = old_row

        self.voice_id_to_row = updated_mappings
        logger.debug(f"重新计算行号映射完成: {self.voice_id_to_row}")

    # ⭐ 新增方法：按行号删除Excel行
    def _delete_excel_row_by_number(self, row_number: int) -> None:
        """
        按行号删除Excel行

        Args:
            row_number: 要删除的行号（从1开始）
        """
        try:
            import openpyxl
            from openpyxl import load_workbook

            workbook = load_workbook(self.filename)
            worksheet = workbook.active

            # 删除指定行
            worksheet.delete_rows(row_number)

            workbook.save(self.filename)
            logger.debug(f"已删除Excel第{row_number}行")

        except Exception as e:
            logger.error(f"删除Excel第{row_number}行失败: {e}")

    # ⭐ 新增方法：写入到指定行
    def _write_to_specific_row(self, row: int, voice_id: int, value: Any, **kwargs) -> None:
        """
        写入数据到指定行

        Args:
            row: Excel行号
            voice_id: 语音录入ID
            value: 测量值
            **kwargs: 其他参数（如原始语音、处理文本等）
        """
        try:
            import openpyxl
            from openpyxl import load_workbook

            workbook = load_workbook(self.filename)
            worksheet = workbook.active

            # 写入数据到指定列
            worksheet.cell(row=row, column=1, value=self.current_standard_id)  # A列: 标准序号
            worksheet.cell(row=row, column=6, value=pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S"))  # F列: 时间戳
            worksheet.cell(row=row, column=7, value=voice_id)  # G列: 语音录入ID

            # 根据数据类型写入测量值
            if isinstance(value, str):
                worksheet.cell(row=row, column=4, value=value)  # D列: 测量值
            else:
                worksheet.cell(row=row, column=4, value=float(value))  # D列: 测量值

            # 写入其他参数
            if 'original_text' in kwargs:
                if '原始语音' in self.columns:
                    col_index = self.columns.index('原始语音') + 1
                    worksheet.cell(row=row, column=col_index, value=kwargs['original_text'])

            if 'processed_text' in kwargs:
                if '处理文本' in self.columns:
                    col_index = self.columns.index('处理文本') + 1
                    worksheet.cell(row=row, column=col_index, value=kwargs['processed_text'])

            workbook.save(self.filename)
            logger.debug(f"数据已写入第{row}行: voice_id={voice_id}, value={value}")

        except Exception as e:
            logger.error(f"写入第{row}行失败: {e}")

    # ⭐ 新增方法：停止时重新编号Excel ID列
    def renumber_excel_ids(self) -> None:
        """
        停止录音时，重新为Excel编号列（C列）分配连续编号

        流程：
        1. 读取当前Excel文件
        2. 获取所有未删除的行（通过语音ID判断）
        3. 按行顺序分配连续编号：1, 2, 3...
        4. 批量更新Excel文件的C列
        """
        try:
            import openpyxl
            from openpyxl import load_workbook

            workbook = load_workbook(self.filename)
            worksheet = workbook.active

            # 获取所有有效记录并按行号排序
            active_records = sorted(self.voice_id_to_row.items(), key=lambda x: x[1])

            if not active_records:
                logger.info("没有需要重新编号的记录")
                return

            # 重新编号
            for i, (voice_id, row) in enumerate(active_records, 1):
                if voice_id not in self.deleted_voice_ids:
                    # 写入Excel编号到C列
                    worksheet.cell(row=row, column=3, value=i)  # C列: Excel编号
                    logger.debug(f"重新编号: voice_id={voice_id}, row={row}, excel_id={i}")

            workbook.save(self.filename)
            logger.info(f"Excel重新编号完成，共{len(active_records)}条记录")

        except Exception as e:
            logger.error(f"Excel重新编号失败: {e}")

    def create_new_file(self) -> None:
        df = pd.DataFrame(columns=self.columns)
        df.to_excel(self.filename, index=False)
        self.format_excel()  # ✅ 仅在此处格式化
        logger.info(f"创建并格式化新Excel文件: {self.filename}")

    def append_with_text(
            self,
            data: List[Tuple[Union[float, str], str, str]],  # (数值或文本, 原始语音文本, 处理文本)
            auto_generate_ids: bool = True
        ) -> List[Tuple[int, Union[float, str], str]]:  # 返回 [(Voice_ID, 数值或文本, 原始文本)]
            """
            📍 修改方法：写入带语音ID的数据（支持双ID系统）
            返回本次写入的所有记录（包含生成的语音ID）
            """
            if not data:
                logger.warning("没有数据可写入")
                return []

            with self._lock:
                try:
                    if not os.path.exists(self.filename):
                        self.create_new_file()

                    # 📍 使用新的写入方式：直接写入到指定行
                    result = []
                    for val, original_text, processed_text in data:
                        # 获取插入位置（语音ID和Excel行号）
                        voice_id, excel_row = self.get_next_insert_position()

                        # 直接写入到指定行
                        self._write_to_specific_row(
                            row=excel_row,
                            voice_id=voice_id,
                            value=val,
                            original_text=original_text,
                            processed_text=processed_text
                        )

                        # 记录到会话数据
                        record_val = val if isinstance(val, str) else self._float_cell(val)
                        self._session_data.append((voice_id, record_val, original_text))

                        # 返回结果
                        result.append((voice_id, record_val, original_text))

                    logger.debug(f"成功写入 {len(result)} 条数据到 {self.filename}")
                    return result

                except Exception as e:
                    logger.error(f"写入Excel失败: {e}")
                    return []

    def append(
        self,
        data: Union[List[float], List[Tuple[int, float]]],
        auto_generate_ids: bool = True
    ) -> bool:
        if not data:
            logger.warning("没有数据可写入")
            return False

        if isinstance(data[0], (int, float)) and auto_generate_ids:
            start_id = self._last_id + 1 if self._last_id > 0 else 1
            data_pairs: List[Tuple[int, float]] = [(start_id + i, self._float_cell(v)) for i, v in enumerate(data)]
        elif isinstance(data[0], (list, tuple)) and len(data[0]) == 2:
            data_pairs = []
            for item in data:
                if isinstance(item, (list, tuple)) and len(item) == 2:
                    data_pairs.append((
                        self._int_cell(item[0]),
                        self._float_cell(item[1])
                    ))
        else:
            logger.error(f"无效的数据格式: {type(data[0])}，期望数值或(编号, 数值)对")
            return False

        return self._write_to_excel_core(data_pairs)

    # Backward compatibility
    def append_rows(self, values: List[float]) -> bool:
        return self.append(values, auto_generate_ids=True)

    def append_to_excel(self, data_pairs: List[Tuple[int, float]]) -> bool:
        return self.append(data_pairs, auto_generate_ids=False)

    def _write_to_excel_core(self, data_pairs: List[Tuple[int, float]]) -> bool:
        if not data_pairs:
            logger.warning("没有数据可写入")
            return False

        with self._lock:
            try:
                if not os.path.exists(self.filename):
                    self.create_new_file()

                existing_data = pd.read_excel(self.filename)
                
                # 获取配置设置
                include_original = config.get("excel.formatting.include_original", True)
                auto_numbering = config.get("excel.formatting.auto_numbering", True)
                include_timestamp = config.get("excel.formatting.include_timestamp", True)
                
                # 生成新数据记录
                new_data = []
                for nid, val in data_pairs:
                    record: Dict[str, Union[int, float, str]] = {}
                    
                    # 根据配置决定是否添加编号
                    if auto_numbering:
                        record["编号"] = nid
                    
                    # 添加测量值
                    record["测量值"] = val
                    
                    # 根据配置决定是否添加时间戳
                    if include_timestamp:
                        record["时间戳"] = pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S")
                    
                    # 根据配置决定是否添加原始语音字段
                    if include_original:
                        record["原始语音"] = ""  # 兼容旧数据
                        
                    new_data.append(record)

                if existing_data.empty:
                    updated_data = pd.DataFrame(new_data)
                else:
                    updated_data = pd.concat([existing_data, pd.DataFrame(new_data)], ignore_index=True)

                updated_data.to_excel(self.filename, index=False)
                  
                if data_pairs and auto_numbering:
                    self._last_id = max(nid for nid, _ in data_pairs)

                logger.info(f"成功写入 {len(data_pairs)} 条数据到 {self.filename}")
                return True

            except Exception as e:
                logger.error(f"写入Excel失败: {e}")
                return False

    def format_excel(self) -> None:
        try:
            workbook = load_workbook(self.filename)
            worksheet = workbook.active
            if worksheet is None:
                return
            
            # 获取include_original配置
            include_original = config.get("excel.formatting.include_original", True)
            
            # 根据配置设置列宽
            column_widths = {'A': 10, 'B': 15, 'C': 20}
            if include_original:
                column_widths['D'] = 30  # 只有在包含原始语音时才设置D列宽
            
            for col_letter, width in column_widths.items():
                worksheet.column_dimensions[col_letter].width = width

            # Format header (row 1)
            header_font = Font(bold=True, size=12)
            header_alignment = Alignment(horizontal='center')
            for col in range(1, len(self.columns) + 1):
                cell = worksheet.cell(row=1, column=col)
                cell.font = header_font
                cell.alignment = header_alignment

            workbook.save(self.filename)
            logger.info("Excel格式化完成")
        except Exception as e:
            logger.warning(f"Excel格式化失败: {e}")

    def get_existing_ids(self) -> List[int]:
        if not os.path.exists(self.filename):
            return []
        try:
            workbook = load_workbook(self.filename, read_only=True)
            worksheet = workbook.active
            if worksheet is None:
                workbook.close()
                return []
            existing_ids: List[int] = []
            for row in range(2, worksheet.max_row + 1):
                cell_value = worksheet.cell(row=row, column=1).value
                if cell_value is not None:
                    try:
                        existing_ids.append(self._int_cell(cell_value))
                    except (ValueError, TypeError):
                        continue
            workbook.close()
            return existing_ids
        except Exception as e:
            logger.error(f"使用openpyxl读取现有编号失败: {e}")
            try:
                df = pd.read_excel(self.filename)
                return df["编号"].tolist() if "编号" in df.columns else []
            except Exception as fallback_error:
                logger.error(f"回退到pandas也失败: {fallback_error}")
                return []

    def get_session_data(self) -> List[Tuple[Union[int, str, float], Any, str]]:
            """获取本次会话的所有数据"""
            return self._session_data.copy()
    
    def clear_session_data(self) -> None:
        """清空会话数据"""
        self._session_data.clear()

# 测试代码（保持不变，但修复日志配置）
if __name__ == '__main__':
    import tempfile, shutil
    logging.basicConfig(level=logging.INFO, format='%(message)s') # ✅ 测试时才配置日志
    test_dir = tempfile.mkdtemp()
    test_file = os.path.join(test_dir, "test_data.xlsx")
    try:
        print("=== 开始优化模块自检 ===")
        exporter = ExcelExporter(filename=test_file)
        test_data: List[Tuple[int, float]] = [(1, 12.5), (2, 33.8), (3, 99.9)]
        print("\n[测试1] 统一接口 - 数据对写入")
        assert exporter.append(test_data, auto_generate_ids=False), "数据对写入失败"
        assert os.path.exists(test_file), "文件未创建"
        print("\n[测试2] 数据读取功能")
        df = pd.read_excel(test_file)
        assert len(df) == 3, "数据量不符"
        assert df['编号'].tolist() == [1, 2, 3], "编号数据错误"
        print("\n[测试3] 统一接口 - 数值列表写入")
        exporter.append([55.5, 77.7, 88.8], auto_generate_ids=True)
        updated_df = pd.read_excel(test_file)
        assert len(updated_df) == 6, "数值列表写入失败"
        print("\n✅ 所有测试通过！")
        print("生成的文件预览:")
        print(updated_df.head())
    except Exception as e:
        print(f"\n❌ 测试失败: {e}")
    finally:
        shutil.rmtree(test_dir)