#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
更新模板文件：调大时间戳列间距，设置所有单元格居中对齐
"""

import os
from openpyxl import load_workbook
from openpyxl.styles import Alignment

def update_template():
    """更新模板文件"""
    print("🔧 更新模板文件")
    print("="*50)

    template_path = "reports/enhanced_measure_template.xlsx"

    if not os.path.exists(template_path):
        print(f"❌ 模板文件不存在: {template_path}")
        return

    # 加载模板
    workbook = load_workbook(template_path)
    worksheet = workbook.active

    print(f"📊 原始模板信息:")
    print(f"   最大行数: {worksheet.max_row}")
    print(f"   最大列数: {worksheet.max_column}")

    # 1. 调大时间戳列间距 (第I列)
    print(f"\n🔧 1. 调大时间戳列间距...")
    worksheet.column_dimensions['I'].width = 25  # 从原来的20调整到25

    # 调整其他列宽，使整体更协调
    column_widths = {
        'A': 12,   # 标准序号
        'B': 15,   # 标准内容
        'C': 12,   # 下限
        'D': 12,   # 上限
        'E': 8,    # 序号
        'F': 12,   # 测量值
        'G': 12,   # 判断结果
        'H': 12,   # 偏差
        'I': 25,   # 时间戳 (调大)
        'J': 15    # 语音录入编号
    }

    for col_letter, width in column_widths.items():
        worksheet.column_dimensions[col_letter].width = width
        print(f"   列 {col_letter}: 宽度 = {width}")

    # 2. 设置所有单元格居中对齐
    print(f"\n🔧 2. 设置所有单元格居中对齐...")
    center_alignment = Alignment(horizontal="center", vertical="center")

    aligned_count = 0
    for row in range(1, worksheet.max_row + 1):
        for col in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=row, column=col)
            if cell.value is not None:  # 只对有内容的单元格设置对齐
                cell.alignment = center_alignment
                aligned_count += 1

    print(f"   已对齐 {aligned_count} 个单元格")

    # 保存更新后的模板
    backup_path = template_path.replace('.xlsx', '_backup.xlsx')
    os.rename(template_path, backup_path)
    print(f"\n📁 原模板已备份为: {backup_path}")

    workbook.save(template_path)
    workbook.close()

    print(f"✅ 模板更新完成: {template_path}")

    # 验证更新结果
    print(f"\n🔍 验证更新结果...")
    workbook = load_workbook(template_path)
    worksheet = workbook.active

    print(f"   时间戳列 (I列) 宽度: {worksheet.column_dimensions['I'].width}")

    # 检查前几行的对齐情况
    print(f"   单元格对齐检查:")
    for row in range(1, min(6, worksheet.max_row + 1)):
        for col in range(1, min(6, worksheet.max_column + 1)):
            cell = worksheet.cell(row=row, column=col)
            if cell.value:
                alignment = cell.alignment
                print(f"     单元格({row},{col}) '{str(cell.value)[:15]}': 水平={alignment.horizontal}, 垂直={alignment.vertical}")

    workbook.close()

    print(f"\n🎉 模板更新完成!")
    print(f"   ✅ 时间戳列间距已调大")
    print(f"   ✅ 所有单元格已设置居中对齐")
    print(f"   ✅ 列宽已优化调整")

if __name__ == "__main__":
    update_template()