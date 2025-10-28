#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
检查自动生成的报告文件
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

def check_generated_report():
    """检查自动生成的报告文件"""
    print("🔍 检查自动生成的报告文件")
    print("="*60)

    file_path = "reports/simple_integration_test.xlsx"

    if not os.path.exists(file_path):
        print(f"❌ 文件不存在: {file_path}")
        return

    # 使用openpyxl读取原始结构
    workbook = load_workbook(file_path)
    worksheet = workbook.active

    print(f"📊 文件信息:")
    print(f"   工作表最大行数: {worksheet.max_row}")
    print(f"   工作表最大列数: {worksheet.max_column}")

    print(f"\n📋 当前文件内容:")
    for row in range(1, min(worksheet.max_row + 1, 12)):
        row_content = []
        for col in range(1, min(worksheet.max_column + 1, 12)):
            cell_value = worksheet.cell(row=row, column=col).value
            if cell_value is not None:
                cell_value = str(cell_value)[:40]
            else:
                cell_value = "None"
            row_content.append(f"{cell_value:>15}")
        print(f"行{row:2d}: {' | '.join(row_content)}")

    # 检查字体样式
    print(f"\n🎨 第1行当前样式:")
    cell = worksheet.cell(row=1, column=1)
    print(f"   字体: {cell.font}")
    print(f"   填充: {cell.fill}")
    print(f"   对齐: {cell.alignment}")

    # 检查重复行
    print(f"\n🔍 检查重复行:")
    row_2_content = []
    row_3_content = []

    for col in range(1, worksheet.max_column + 1):
        row_2_content.append(str(worksheet.cell(row=2, column=col).value or ""))
        row_3_content.append(str(worksheet.cell(row=3, column=col).value or ""))

    print(f"   第2行: {row_2_content}")
    print(f"   第3行: {row_3_content}")

    if row_2_content == row_3_content:
        print("   ✅ 发现重复行，第2行将被删除，第3行保留")
    else:
        print("   ❌ 第2行和第3行内容不同")

    workbook.close()

if __name__ == "__main__":
    import os
    check_generated_report()