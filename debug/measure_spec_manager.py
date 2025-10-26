#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
测量规范管理器
自动查询测量规范并更新Excel报告
"""

import os
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from typing import Dict, Any, Optional, Tuple
import logging

logger = logging.getLogger(__name__)

class MeasureSpecManager:
    """测量规范管理器"""

    def __init__(self, spec_directory: str = "reports"):
        self.spec_directory = spec_directory
        self.spec_cache = {}  # 缓存已加载的规范文件

    def load_measure_spec(self, part_no: str) -> Optional[pd.DataFrame]:
        """
        加载测量规范文件

        Args:
            part_no: 零件号

        Returns:
            测量规范DataFrame，如果文件不存在返回None
        """
        # 检查缓存
        if part_no in self.spec_cache:
            return self.spec_cache[part_no]

        # 构建文件名
        spec_filename = f"{part_no}_MeasureSpec.xlsx"
        spec_filepath = os.path.join(self.spec_directory, spec_filename)

        if not os.path.exists(spec_filepath):
            logger.warning(f"测量规范文件不存在: {spec_filepath}")
            return None

        try:
            # 使用openpyxl读取Excel文件，跳过前2行
            workbook = load_workbook(spec_filepath)
            worksheet = workbook.active

            # 从第3行开始读取数据（跳过标题行）
            data = []
            for row in range(3, worksheet.max_row + 1):
                row_data = []
                for col in range(1, worksheet.max_column + 1):
                    cell_value = worksheet.cell(row=row, column=col).value
                    row_data.append(cell_value)

                # 如果行不为空，添加到数据中
                if any(cell is not None for cell in row_data):
                    data.append(row_data)

            workbook.close()

            if not data:
                logger.warning(f"测量规范文件没有有效数据: {spec_filename}")
                return None

            # 创建DataFrame
            df = pd.DataFrame(data)

            # 重命名列，使其更有意义
            if len(df.columns) >= 5:
                df.columns = ['标准序号', '标准内容', '下限', '上限', '备注']
            elif len(df.columns) >= 4:
                df.columns = ['标准序号', '标准内容', '下限', '上限']
            else:
                logger.error(f"测量规范文件格式不正确: {spec_filename}")
                return None

            # 过滤出有效的数据行（标准序号为数字的行）
            valid_rows = []
            for _, row in df.iterrows():
                try:
                    standard_id = row['标准序号']
                    if pd.notna(standard_id) and str(standard_id).isdigit():
                        valid_rows.append(row)
                except:
                    continue

            if not valid_rows:
                logger.warning(f"测量规范文件没有有效的标准序号数据: {spec_filename}")
                return None

            df = pd.DataFrame(valid_rows)

            # 转换数据类型
            df['标准序号'] = pd.to_numeric(df['标准序号'])
            df['下限'] = pd.to_numeric(df['下限'], errors='coerce')
            df['上限'] = pd.to_numeric(df['上限'], errors='coerce')

            # 缓存结果
            self.spec_cache[part_no] = df

            logger.info(f"成功加载测量规范: {spec_filename}, 共{len(df)}条记录")
            return df

        except Exception as e:
            logger.error(f"加载测量规范文件失败: {e}")
            import traceback
            traceback.print_exc()
            return None

    def get_measure_spec(self, part_no: str, standard_id: int) -> Optional[Dict[str, Any]]:
        """
        获取指定零件号和标准序号的测量规范

        Args:
            part_no: 零件号
            standard_id: 标准序号

        Returns:
            测量规范字典，包含下限、上限、标准内容等
        """
        df = self.load_measure_spec(part_no)
        if df is None:
            return None

        # 查找匹配的记录
        matching_rows = df[df['标准序号'] == standard_id]

        if matching_rows.empty:
            logger.warning(f"未找到标准序号 {standard_id} 的测量规范")
            return None

        # 取第一个匹配记录
        spec_row = matching_rows.iloc[0]

        return {
            'standard_id': int(spec_row['标准序号']),
            'content': spec_row['标准内容'],
            'lower_limit': spec_row['下限'],
            'upper_limit': spec_row['上限'],
            'remark': spec_row.get('备注', '')
        }

    def judge_measurement(self, part_no: str, standard_id: int, measured_value: float) -> Dict[str, Any]:
        """
        判断测量结果

        Args:
            part_no: 零件号
            standard_id: 标准序号
            measured_value: 测量值

        Returns:
            判断结果字典
        """
        spec = self.get_measure_spec(part_no, standard_id)
        if spec is None:
            return {
                'standard_id': standard_id,
                'content': '未知',
                'lower_limit': None,
                'upper_limit': None,
                'measured_value': measured_value,
                'result': '无规范',
                'is_ok': None,
                'deviation': None
            }

        lower_limit = spec['lower_limit']
        upper_limit = spec['upper_limit']

        # 判断结果
        if pd.isna(lower_limit) and pd.isna(upper_limit):
            result = '无规范'
            is_ok = None
            deviation = None
        elif pd.isna(lower_limit):  # 只有上限
            if measured_value <= upper_limit:
                result = 'OK'
                is_ok = True
                deviation = upper_limit - measured_value
            else:
                result = 'NOK'
                is_ok = False
                deviation = measured_value - upper_limit
        elif pd.isna(upper_limit):  # 只有下限
            if measured_value >= lower_limit:
                result = 'OK'
                is_ok = True
                deviation = measured_value - lower_limit
            else:
                result = 'NOK'
                is_ok = False
                deviation = lower_limit - measured_value
        else:  # 既有上限又有下限
            if lower_limit <= measured_value <= upper_limit:
                result = 'OK'
                is_ok = True
                # 计算到最近的边界的偏差
                deviation_to_lower = measured_value - lower_limit
                deviation_to_upper = upper_limit - measured_value
                deviation = min(deviation_to_lower, deviation_to_upper)
            else:
                result = 'NOK'
                is_ok = False
                # 计算超出范围的偏差
                if measured_value < lower_limit:
                    deviation = lower_limit - measured_value
                else:
                    deviation = measured_value - upper_limit

        return {
            'standard_id': standard_id,
            'content': spec['content'],
            'lower_limit': lower_limit,
            'upper_limit': upper_limit,
            'measured_value': measured_value,
            'result': result,
            'is_ok': is_ok,
            'deviation': deviation
        }

    def update_excel_with_specs(self, excel_filepath: str, part_no: str) -> bool:
        """
        更新Excel文件，添加测量规范和判断结果

        Args:
            excel_filepath: Excel文件路径
            part_no: 零件号

        Returns:
            是否更新成功
        """
        if not os.path.exists(excel_filepath):
            logger.error(f"Excel文件不存在: {excel_filepath}")
            return False

        try:
            # 使用openpyxl打开Excel文件
            workbook = load_workbook(excel_filepath)
            worksheet = workbook.active

            # 找到数据开始行（通常是第4行）
            data_start_row = 4

            # 获取列索引
            col_mapping = self._get_column_mapping(worksheet, data_start_row)

            # 更新每一行数据
            for row in range(data_start_row, worksheet.max_row + 1):
                # 获取标准序号
                standard_id_cell = worksheet.cell(row=row, column=col_mapping['standard_id'])
                if standard_id_cell.value is None:
                    continue

                try:
                    standard_id = int(standard_id_cell.value)
                except (ValueError, TypeError):
                    continue

                # 获取测量值
                measured_value_cell = worksheet.cell(row=row, column=col_mapping['measured_value'])
                if measured_value_cell.value is None:
                    continue

                try:
                    measured_value = float(measured_value_cell.value)
                except (ValueError, TypeError):
                    continue

                # 获取测量规范和判断结果
                judgment = self.judge_measurement(part_no, standard_id, measured_value)

                # 更新标准内容
                if 'content' in col_mapping and col_mapping['content']:
                    content_cell = worksheet.cell(row=row, column=col_mapping['content'])
                    content_cell.value = judgment['content']

                # 更新下限
                if 'lower_limit' in col_mapping and col_mapping['lower_limit']:
                    lower_limit_cell = worksheet.cell(row=row, column=col_mapping['lower_limit'])
                    lower_limit_cell.value = judgment['lower_limit']

                # 更新上限
                if 'upper_limit' in col_mapping and col_mapping['upper_limit']:
                    upper_limit_cell = worksheet.cell(row=row, column=col_mapping['upper_limit'])
                    upper_limit_cell.value = judgment['upper_limit']

                # 更新判断结果
                if 'judgment' in col_mapping and col_mapping['judgment']:
                    judgment_cell = worksheet.cell(row=row, column=col_mapping['judgment'])
                    judgment_cell.value = judgment['result']

                    # 设置单元格格式
                    if judgment['is_ok'] is True:
                        judgment_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # 绿色
                    elif judgment['is_ok'] is False:
                        judgment_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # 红色

                # 更新偏差
                if 'deviation' in col_mapping and col_mapping['deviation'] and judgment['deviation'] is not None:
                    deviation_cell = worksheet.cell(row=row, column=col_mapping['deviation'])
                    deviation_cell.value = round(judgment['deviation'], 2)

            # 保存文件
            workbook.save(excel_filepath)
            workbook.close()

            logger.info(f"成功更新Excel文件: {excel_filepath}")
            return True

        except Exception as e:
            logger.error(f"更新Excel文件失败: {e}")
            return False

    def _get_column_mapping(self, worksheet, header_row: int) -> Dict[str, int]:
        """
        获取列映射

        Args:
            worksheet: 工作表
            header_row: 标题行号

        Returns:
            列名到列索引的映射
        """
        mapping = {
            'standard_id': None,
            'content': None,
            'lower_limit': None,
            'upper_limit': None,
            'measured_value': None,
            'judgment': None,
            'deviation': None
        }

        # 读取标题行
        for col in range(1, worksheet.max_column + 1):
            cell_value = worksheet.cell(row=header_row, column=col).value
            if cell_value is None:
                continue

            cell_value = str(cell_value).strip()

            # 根据列名映射
            if '标准序号' in cell_value or '标准' in cell_value:
                mapping['standard_id'] = col
            elif '标准内容' in cell_value or '内容' in cell_value:
                mapping['content'] = col
            elif '下限' in cell_value:
                mapping['lower_limit'] = col
            elif '上限' in cell_value:
                mapping['upper_limit'] = col
            elif '测量值' in cell_value:
                mapping['measured_value'] = col
            elif '判断结果' in cell_value or '结果' in cell_value:
                mapping['judgment'] = col
            elif '偏差' in cell_value:
                mapping['deviation'] = col

        return mapping

    def clear_cache(self):
        """清空缓存"""
        self.spec_cache.clear()

# 测试代码
if __name__ == "__main__":
    # 设置日志
    logging.basicConfig(level=logging.INFO)

    # 创建管理器
    manager = MeasureSpecManager()

    # 测试加载规范
    print("测试加载测量规范...")
    spec = manager.load_measure_spec("PART-A001")
    if spec is not None:
        print(f"加载成功，共{len(spec)}条记录")
        print(spec.head())

    # 测试判断测量结果
    print("\n测试判断测量结果...")
    judgment = manager.judge_measurement("PART-A001", 100, 80.0)
    print(f"判断结果: {judgment}")

    # 测试更新Excel文件
    print("\n测试更新Excel文件...")
    excel_file = "reports/Report_PART-A001_BATCH-202501_20251026_140920.xlsx"
    success = manager.update_excel_with_specs(excel_file, "PART-A001")
    print(f"更新结果: {'成功' if success else '失败'}")