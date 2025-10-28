#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
测试修复后的报告生成功能
"""

import os
import sys
import pandas as pd

# 添加项目根目录到Python路径
project_root = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, project_root)

from excel_exporter_enhanced import ExcelExporterEnhanced

def test_fixed_report():
    """测试修复后的报告生成"""
    print("🎯 测试修复后的报告生成功能")
    print("="*60)

    # 创建测量规范文件
    spec_data = [
        [100, '半径1', 75.5, 85.8],
        [200, '半径2', 15.5, 30.5],
        [300, '半径3', 8.5, 12.5],
    ]

    spec_df = pd.DataFrame(spec_data, columns=['标准序号', '标准内容', '下限', '上限'])
    spec_filename = "reports/FIXED-TEST_MeasureSpec.xlsx"

    os.makedirs("reports", exist_ok=True)
    spec_df.to_excel(spec_filename, index=False)

    # 创建导出器
    exporter = ExcelExporterEnhanced(
        filename="reports/fixed_report_test.xlsx",
        part_no="FIXED-TEST",
        batch_no="B001",
        inspector="修复测试员"
    )

    print("📝 1. 使用模板创建文件...")
    success = exporter.create_from_template("FIXED-TEST", "B001", "修复测试员")
    print(f"   模板创建结果: {success}")

    print("📝 2. 写入测试数据...")
    test_scenarios = [
        (100, 80.0, "半径1 八十"),    # OK
        (200, 25.0, "半径2 二十五"),  # OK
        (300, 10.0, "半径3 十"),      # NOK
    ]

    for standard_id, value, text in test_scenarios:
        exporter.current_standard_id = standard_id
        test_data = [(value, text, text)]
        results = exporter.append_with_text(test_data)
        print(f"   标准序号{standard_id}: 值{value} -> 写入结果: {results}")

    print("🔧 3. 执行最终格式化...")
    success = exporter.finalize_excel_file()
    print(f"   格式化结果: {success}")

    if success and os.path.exists(exporter.filename):
        print(f"\n📋 4. 验证结果:")

        # 使用openpyxl检查样式
        from openpyxl import load_workbook
        workbook = load_workbook(exporter.filename)
        worksheet = workbook.active

        print(f"   Excel文件包含 {worksheet.max_row} 行数据")

        # 检查第一行样式
        title_cell = worksheet.cell(row=1, column=1)
        print(f"\n🎨 第一行样式检查:")
        print(f"   内容: {title_cell.value}")
        print(f"   字体颜色: {title_cell.font.color}")
        print(f"   字体大小: {title_cell.font.sz}")
        print(f"   是否加粗: {title_cell.font.b}")
        print(f"   对齐方式: 水平={title_cell.alignment.horizontal}, 垂直={title_cell.alignment.vertical}")

        # 检查前几行内容
        print(f"\n📋 文件前5行内容:")
        for row in range(1, min(6, worksheet.max_row + 1)):
            row_content = []
            for col in range(1, min(11, worksheet.max_column + 1)):
                cell_value = worksheet.cell(row=row, column=col).value
                if cell_value is not None:
                    cell_value = str(cell_value)[:30]
                else:
                    cell_value = "None"
                row_content.append(f"{cell_value:>15}")
            print(f"   行{row:2d}: {' | '.join(row_content)}")

        workbook.close()

        # 清理文件
        os.remove(exporter.filename)
        print(f"\n🧹 已清理测试文件")

    # 清理测量规范文件
    if os.path.exists(spec_filename):
        os.remove(spec_filename)
        print(f"🧹 已清理测量规范文件")

if __name__ == "__main__":
    test_fixed_report()