#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
创建增强的Excel模板
包含测量规范和自动判断功能
"""

import os
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

def create_enhanced_template():
    """创建增强的Excel模板"""

    # 模板文件路径
    template_path = "reports/enhanced_report_template.xlsx"

    # 确保目录存在
    os.makedirs("reports", exist_ok=True)

    # 创建数据
    data = {
        '标准序号': [''],
        '标准内容': [''],
        '下限': [''],
        '上限': [''],
        '测量值': [''],
        '判断结果': [''],
        '偏差': [''],
        '时间戳': [''],
        '语音录入编号': ['']
    }

    df = pd.DataFrame(data)

    # 保存为Excel
    df.to_excel(template_path, index=False)

    # 使用openpyxl进行格式化
    workbook = load_workbook(template_path)
    worksheet = workbook.active

    # 设置列宽
    column_widths = {
        'A': 10,  # 标准序号
        'B': 20,  # 标准内容
        'C': 12,  # 下限
        'D': 12,  # 上限
        'E': 12,  # 测量值
        'F': 12,  # 判断结果
        'G': 12,  # 偏差
        'H': 20,  # 时间戳
        'I': 15   # 语音录入编号
    }

    for col_letter, width in column_widths.items():
        worksheet.column_dimensions[col_letter].width = width

    # 格式化标题行
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    for col in range(1, len(data.keys()) + 1):
        cell = worksheet.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # 添加边框
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 为数据区域添加边框
    for row in range(1, worksheet.max_row + 1):
        for col in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.border = thin_border

    # 在第3行添加说明信息
    info_row = 3
    info_data = [
        ("说明", "", "", "", "", "", "", "", ""),
        ("1. 标准序号: 对应测量规范中的标准序号", "", "", "", "", "", "", "", ""),
        ("2. 标准内容: 自动从测量规范文件中获取", "", "", "", "", "", "", "", ""),
        ("3. 下限/上限: 自动从测量规范文件中获取", "", "", "", "", "", "", "", ""),
        ("4. 测量值: 语音识别的实际测量结果", "", "", "", "", "", "", "", ""),
        ("5. 判断结果: 自动根据测量规范判断OK/NOK", "", "", "", "", "", "", "", ""),
        ("6. 偏差: 测量值与规范边界的差值", "", "", "", "", "", "", "", "")
    ]

    for col, (text, *_) in enumerate(info_data, 1):
        cell = worksheet.cell(row=info_row, column=col)
        cell.value = text
        cell.font = Font(size=10, italic=True)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        if col == 1:
            cell.font = Font(size=10, bold=True, italic=True)

    # 合并说明行的单元格
    for row_idx, row_data in enumerate(info_data, info_row):
        if row_idx > info_row:  # 跳过标题行
            worksheet.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=9)

    # 设置数据开始行为第10行（跳过说明信息）
    data_start_row = 10

    # 添加一些示例公式（可选）
    # 这些公式可以在实际使用时启用

    # 在数据开始行添加示例（这些会在实际使用时被覆盖）
    example_row = data_start_row
    worksheet.cell(row=example_row, column=1, value=100)  # 标准序号示例
    worksheet.cell(row=example_row, column=5, value=0)    # 测量值示例（初始为0）

    # 添加判断结果的示例公式（注释掉，实际使用时由Python代码处理）
    # judgment_formula = '=IF(AND(E2>=C2, E2<=D2), "OK", "NOK")'
    # worksheet.cell(row=example_row, column=6, value=judgment_formula)

    # 保存模板
    workbook.save(template_path)
    workbook.close()

    print(f"✅ 增强Excel模板已创建: {template_path}")
    print(f"📋 模板包含列: {list(data.keys())}")
    print(f"📝 数据开始行: 第{data_start_row}行")

    return template_path

if __name__ == "__main__":
    create_enhanced_template()