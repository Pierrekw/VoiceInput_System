#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
测试Excel测量规范文件缺失时的警告功能
"""

import sys
import os

# 添加父目录到路径
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

def test_measure_spec_warning():
    """测试测量规范文件缺失时的警告功能"""
    print("🧪 测试Excel测量规范文件缺失警告功能")
    print("-" * 50)

    try:
        from excel_utils import ExcelExporterEnhanced
        import logging

        # 设置详细日志
        logging.basicConfig(level=logging.DEBUG)

        # 创建Excel导出器，使用不存在的零件号
        test_filename = "test_measure_spec_warning.xlsx"
        nonexistent_part_no = "PART-999"  # 不存在的零件号
        exporter = ExcelExporterEnhanced(filename=test_filename, part_no=nonexistent_part_no)

        # 创建模板
        success = exporter.create_from_template(nonexistent_part_no, "B202501", "测试员")
        if success:
            print("✅ Excel模板创建成功")
        else:
            print("❌ Excel模板创建失败")
            return False

        # 添加一些测试数据
        test_data = [
            (25.4, "二十五点四", "25.4"),  # 标准序号100
            (30.1, "三十点一", "30.1"),   # 标准序号200
            (28.7, "二十八点七", "28.7")  # 标准序号300
        ]

        # 设置不同的标准序号并添加数据
        for i, (val, original, processed) in enumerate(test_data):
            exporter.current_standard_id = [100, 200, 300][i]
            results = exporter.append_with_text([(val, original, processed)])
            print(f"✅ 添加数据: 标准序号={exporter.current_standard_id}, 值={val}")

        print(f"总共添加了 {len(test_data)} 条数据")

        # 执行最终格式化
        print("🔧 开始执行Excel最终格式化（预期会显示测量规范文件缺失警告）...")
        success = exporter.finalize_excel_file()

        if success:
            print("✅ Excel最终格式化成功")

            # 验证结果
            import openpyxl
            workbook = openpyxl.load_workbook(test_filename)
            worksheet = workbook.active

            print("\n📊 验证Excel内容:")
            found_warning = False
            found_data = False

            for row in range(5, worksheet.max_row + 1):  # 从第5行开始是数据
                standard_id = worksheet.cell(row=row, column=1).value       # 标准序号
                warning_msg = worksheet.cell(row=row, column=2).value      # 警告信息（可能在第2列）
                measured_value = worksheet.cell(row=row, column=6).value    # 测量值
                voice_id = worksheet.cell(row=row, column=10).value         # Voice ID

                print(f"第{row}行: 标准序号={standard_id}, 测量值={measured_value}, Voice ID={voice_id}")

                if warning_msg and "警告" in str(warning_msg):
                    print(f"  ⚠️ 警告信息: {warning_msg}")
                    found_warning = True

                if measured_value is not None and standard_id is not None:
                    found_data = True
                    print(f"  ✅ 历史数据已保留")

            workbook.close()

            if found_warning and found_data:
                print("\n✅ 测量规范文件缺失警告功能正常工作！")
                print("  - 显示了警告信息")
                print("  - 保留了历史识别数据")
                return True
            else:
                print(f"\n❌ 警告功能异常: 找到警告={found_warning}, 找到数据={found_data}")
                return False
        else:
            print("❌ Excel最终格式化失败")
            return False

    except Exception as e:
        print(f"❌ 测试失败: {e}")
        import traceback
        traceback.print_exc()
        return False

    finally:
        # 清理测试文件
        if os.path.exists("test_measure_spec_warning.xlsx"):
            os.remove("test_measure_spec_warning.xlsx")

if __name__ == "__main__":
    success = test_measure_spec_warning()
    if success:
        print("\n🎉 测量规范文件缺失警告功能测试成功！")
    else:
        print("\n⚠️ 测量规范文件缺失警告功能测试失败。")
    sys.exit(0 if success else 1)