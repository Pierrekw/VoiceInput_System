#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
检查新模板文件的结构
"""

import os
from openpyxl import load_workbook

def check_template():
    """检查模板文件结构"""
    print("🔍 检查模板文件结构")
    print("="*50)

    template_path = "reports/templates/enhanced_measure_template.xlsx"

    if not os.path.exists(template_path):
        print(f"❌ 模板文件不存在: {template_path}")
        return

    workbook = load_workbook(template_path)
    worksheet = workbook.active

    print(f"📊 模板信息:")
    print(f"   最大行数: {worksheet.max_row}")
    print(f"   最大列数: {worksheet.max_column}")

    print(f"\n📋 模板内容:")
    for row in range(1, min(worksheet.max_row + 1, 10)):
        row_content = []
        for col in range(1, min(worksheet.max_column + 1, 12)):
            cell_value = worksheet.cell(row=row, column=col).value
            if cell_value is not None:
                cell_value = str(cell_value)[:30]
            else:
                cell_value = "None"
            row_content.append(f"{cell_value:>15}")
        print(f"行{row:2d}: {' | '.join(row_content)}")

    # 检查合并单元格
    print(f"\n🔗 合并单元格检查:")
    if hasattr(worksheet, 'merged_cells') and worksheet.merged_cells:
        for merged_range in worksheet.merged_cells.ranges:
            print(f"   合并范围: {merged_range}")
    else:
        print(f"   无合并单元格")

    workbook.close()

if __name__ == "__main__":
    check_template()
