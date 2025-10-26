#!/usr/bin/env python3
"""
PyInstaller运行时钩子 - 处理模型路径和依赖检查
"""

import os
import sys
import site
from pathlib import Path

def check_and_download_models():
    """检查模型文件是否存在，如果不存在则提示用户"""
    # 检查FunASR模型
    model_path = Path("model/fun")
    if not model_path.exists():
        print("=" * 60)
        print("⚠️  模型文件缺失！")
        print("=" * 60)
        print("模型文件未找到，请按以下步骤操作：")
        print("1. 下载模型文件包")
        print("2. 将 model/fun 文件夹放置在程序同目录下")
        print("3. 重新运行程序")
        print("=" * 60)

        # 创建模型目录结构
        model_path.mkdir(parents=True, exist_ok=True)

        # 创建模型文件说明
        readme_path = model_path / "README.txt"
        with open(readme_path, 'w', encoding='utf-8') as f:
            f.write("""FunASR模型文件目录
=================

请将以下文件放置在此目录：
- damo/speech_asr_nat-zh-cn_16k-common-vocab8484-pytorch.model
- damo/speech_asr_nat-zh-cn_16k-common-vocab8484-pytorch.yaml
- damo/fsmn_vad_common-zh-cn-16k-common-pytorch
- damo/speech_separation_noh_16k_1684_snapshot.onnx
- damo/speech_timestamp_prediction-16k-common-zh-cn-2024-03-14.model.onnx

这些文件可以从以下位置获取：
- 原始开发环境
- 模型文件包（单独提供）
- 官方FunASR仓库
""")

        input("按回车键退出...")
        sys.exit(1)

def check_onnx_deps():
    """检查ONNX依赖"""
    onnx_path = Path("onnx_deps")
    if not onnx_path.exists():
        print("⚠️ onnx_deps目录不存在，但程序仍可运行")
        # 创建目录
        onnx_path.mkdir(exist_ok=True)

def check_templates():
    """检查Excel模板文件"""
    template_path = Path("reports/templates")
    if not template_path.exists():
        print("📊 创建Excel模板目录...")
        template_path.mkdir(parents=True, exist_ok=True)

        # 如果模板文件不存在，创建一个基本的模板
        template_file = template_path / "enhanced_measure_template.xlsx"
        if not template_file.exists():
            print("📝 创建基本Excel模板...")
            try:
                import openpyxl
                from openpyxl import Workbook

                wb = Workbook()
                ws = wb.active
                ws.title = "测量报告"

                # 基本表头
                headers = ["标准序号", "标准内容", "下限", "上限", "测量值", "判断结果", "偏差", "语音录入编号", "时间戳"]
                for col, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col, value=header)

                # 设置列宽
                for col in range(1, 10):
                    ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15

                wb.save(template_file)
                print(f"✅ Excel模板已创建: {template_file}")
            except Exception as e:
                print(f"⚠️ 创建Excel模板失败: {e}")

def main():
    """主函数"""
    print("🔍 检查运行环境...")

    # 检查模型文件
    check_and_download_models()

    # 检查ONNX依赖
    check_onnx_deps()

    # 检查模板文件
    check_templates()

    # 添加当前目录到Python路径（用于导入本地模块）
    current_dir = Path(__file__).parent
    if str(current_dir) not in sys.path:
        sys.path.insert(0, str(current_dir))

    print("✅ 环境检查完成")

if __name__ == "__main__":
    main()