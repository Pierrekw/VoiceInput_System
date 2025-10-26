#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
检查Excel文件内容
"""

import pandas as pd
from openpyxl import load_workbook

def check_excel_file(filepath):
    """检查Excel文件内容"""
    print(f"📋 检查Excel文件: {filepath}")
    print("="*80)

    # 使用pandas读取
    try:
        df = pd.read_excel(filepath)
        print(f"\n📊 pandas读取结果 (共{len(df)}行):")
        print("-" * 80)
        for idx, row in df.iterrows():
            print(f"行{idx+1}: {dict(row)}")
    except Exception as e:
        print(f"❌ pandas读取失败: {e}")

    print("\n" + "="*80)

    # 使用openpyxl读取
    try:
        workbook = load_workbook(filepath)
        worksheet = workbook.active

        print(f"\n📖 openpyxl读取结果:")
        print(f"工作表最大行数: {worksheet.max_row}")
        print(f"工作表最大列数: {worksheet.max_column}")

        print("\n📋 单元格内容:")
        for row in range(1, min(worksheet.max_row + 1, 15)):  # 只显示前15行
            row_content = []
            for col in range(1, min(worksheet.max_column + 1, 10)):  # 只显示前10列
                cell_value = worksheet.cell(row=row, column=col).value
                if cell_value is not None:
                    cell_value = str(cell_value)[:30]  # 限制显示长度
                else:
                    cell_value = "None"
                row_content.append(f"{cell_value:>15}")
            print(f"行{row:2d}: {' | '.join(row_content)}")

        workbook.close()
    except Exception as e:
        print(f"❌ openpyxl读取失败: {e}")

if __name__ == "__main__":
    # 检查问题文件
    check_excel_file("reports/Report_FINAL-TEST_B202501_20251026_150232.xlsx")