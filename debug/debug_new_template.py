#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
调试新模板的使用情况
"""

import os
import sys
import pandas as pd
from openpyxl import load_workbook

# 添加项目根目录到Python路径
project_root = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, project_root)

from excel_exporter_enhanced import ExcelExporterEnhanced

def debug_new_template():
    """调试新模板的使用"""
    print("🔍 调试新模板的使用")
    print("="*60)

    # 创建测量规范文件
    spec_data = [
        [100, '半径1', 75.5, 85.8],
        [200, '半径2', 15.5, 30.5],
    ]

    spec_df = pd.DataFrame(spec_data, columns=['标准序号', '标准内容', '下限', '上限'])
    spec_filename = "reports/debug_MeasureSpec.xlsx"

    os.makedirs("reports", exist_ok=True)
    spec_df.to_excel(spec_filename, index=False)

    # 创建导出器，使用模板
    exporter = ExcelExporterEnhanced(
        filename="debug_template_test.xlsx",
        part_no="DEBUG-TEST",
        batch_no="B001",
        inspector="调试员"
    )

    print("📋 1. 使用模板创建文件...")
    success = exporter.create_from_template("DEBUG-TEST", "B001", "调试员")
    print(f"   模板创建结果: {success}")

    if os.path.exists(exporter.filename):
        print(f"\n📖 2. 检查创建后的文件结构:")
        workbook = load_workbook(exporter.filename)
        worksheet = workbook.active

        print(f"   工作表最大行数: {worksheet.max_row}")
        print(f"   工作表最大列数: {worksheet.max_column}")

        print("\n   文件内容:")
        for row in range(1, min(worksheet.max_row + 1, 10)):
            row_content = []
            for col in range(1, min(worksheet.max_column + 1, 10)):
                cell_value = worksheet.cell(row=row, column=col).value
                if cell_value is not None:
                    cell_value = str(cell_value)[:30]
                else:
                    cell_value = "None"
                row_content.append(f"{cell_value:>12}")
            print(f"   行{row:2d}: {' | '.join(row_content)}")

        workbook.close()

    print("\n📝 3. 写入测试数据...")
    test_data = [(80.0, "测试数据", "测试数据")]
    results = exporter.append_with_text(test_data)
    print(f"   写入结果: {results}")

    if os.path.exists(exporter.filename):
        print(f"\n📖 4. 写入数据后的文件结构:")
        workbook = load_workbook(exporter.filename)
        worksheet = workbook.active

        print(f"   工作表最大行数: {worksheet.max_row}")
        print(f"   工作表最大列数: {worksheet.max_column}")

        print("\n   文件内容:")
        for row in range(1, min(worksheet.max_row + 1, 15)):
            row_content = []
            for col in range(1, min(worksheet.max_column + 1, 10)):
                cell_value = worksheet.cell(row=row, column=col).value
                if cell_value is not None:
                    cell_value = str(cell_value)[:30]
                else:
                    cell_value = "None"
                row_content.append(f"{cell_value:>12}")
            print(f"   行{row:2d}: {' | '.join(row_content)}")

        # 测试查找数据开始行
        print(f"\n🔍 5. 测试查找数据开始行:")
        data_start_row = exporter._find_data_start_row(worksheet)
        print(f"   找到的数据开始行: {data_start_row}")

        workbook.close()

    # 清理文件
    if os.path.exists(exporter.filename):
        os.remove(exporter.filename)
    if os.path.exists(spec_filename):
        os.remove(spec_filename)

    print(f"\n🧹 清理完成")

if __name__ == "__main__":
    debug_new_template()