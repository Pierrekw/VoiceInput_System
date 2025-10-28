#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
检查Excel文件的格式和内容
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

def check_excel_format():
    """检查Excel文件格式"""
    print("🔍 检查Excel文件格式和内容")
    print("="*60)

    file_path = "reports/test20rows_auto_format.xlsx"

    # 使用pandas读取内容
    try:
        df = pd.read_excel(file_path)
        print(f"📊 Pandas读取结果:")
        print(f"   数据形状: {df.shape}")
        print(f"   列名: {list(df.columns)}")
        print(f"\\n📋 前5行数据:")
        for i, row in df.head().iterrows():
            print(f"   行{i+1}: {dict(row)}")
    except Exception as e:
        print(f"❌ Pandas读取失败: {e}")

    # 使用openpyxl读取原始结构
    workbook = load_workbook(file_path)
    worksheet = workbook.active

    print(f"\\n📊 OpenPyXL读取结果:")
    print(f"   工作表最大行数: {worksheet.max_row}")
    print(f"   工作表最大列数: {worksheet.max_column}")

    print(f"\\n📋 完整文件内容:")
    for row in range(1, min(worksheet.max_row + 1, 10)):
        row_content = []
        for col in range(1, min(worksheet.max_column + 1, 12)):
            cell_value = worksheet.cell(row=row, column=col).value
            if cell_value is not None:
                cell_value = str(cell_value)[:40]
            else:
                cell_value = "None"
            row_content.append(f"{cell_value:>15}")
        print(f"行{row:2d}: {' | '.join(row_content)}")

    # 检查字体和对齐
    print(f"\\n🎨 格式检查:")
    for row in range(1, min(6, worksheet.max_row + 1)):
        for col in range(1, min(6, worksheet.max_column + 1)):
            cell = worksheet.cell(row=row, column=col)
            if cell.value:
                font_info = f"字体大小:{cell.font.sz}, 颜色:{cell.font.color}, 加粗:{cell.font.b}"
                align_info = f"水平:{cell.alignment.horizontal}, 垂直:{cell.alignment.vertical}"
                print(f"   单元格({row},{col}): '{cell.value}' - {font_info}, {align_info}")

    workbook.close()

if __name__ == "__main__":
    check_excel_format()