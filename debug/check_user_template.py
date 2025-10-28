#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
检查用户更新的模板结构
"""

import pandas as pd
from openpyxl import load_workbook

def check_user_template():
    """检查用户模板的精确结构"""
    print("🔍 检查用户更新的模板结构")
    print("="*60)

    template_path = "reports/enhanced_measure_template.xlsx"

    # 使用pandas读取
    try:
        df = pd.read_excel(template_path, header=None)
        print(f"\n📊 pandas读取结果 (共{len(df)}行):")
        print("-" * 80)
        for idx, row in df.iterrows():
            print(f"行{idx+1}: {list(row)}")
    except Exception as e:
        print(f"❌ pandas读取失败: {e}")

    print("\n" + "="*80)

    # 使用openpyxl读取原始单元格值
    try:
        workbook = load_workbook(template_path)
        worksheet = workbook.active

        print(f"\n📖 openpyxl原始单元格值:")
        print(f"工作表最大行数: {worksheet.max_row}")
        print(f"工作表最大列数: {worksheet.max_column}")

        print("\n📋 完整单元格内容:")
        for row in range(1, min(worksheet.max_row + 1, 15)):
            row_content = []
            for col in range(1, min(worksheet.max_column + 1, 12)):
                cell_value = worksheet.cell(row=row, column=col).value
                if cell_value is not None:
                    cell_value = str(cell_value)[:50]  # 限制显示长度
                else:
                    cell_value = "None"
                row_content.append(f"{cell_value:>15}")
            print(f"行{row:2d}: {' | '.join(row_content)}")

        # 检查合并单元格
        if hasattr(worksheet, 'merged_cells'):
            print(f"\n🔗 合并单元格: {worksheet.merged_cells.ranges}")

        workbook.close()
    except Exception as e:
        print(f"❌ openpyxl读取失败: {e}")

if __name__ == "__main__":
    check_user_template()