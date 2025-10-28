#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
单独测试Excel测量规范查询功能
"""

import sys
import os

# 添加父目录到路径
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

def test_measure_spec_query():
    """测试测量规范查询功能"""
    print("🧪 测试Excel测量规范查询功能")
    print("-" * 50)

    try:
        from excel_utils import ExcelExporterEnhanced
        import logging

        # 设置详细日志
        logging.basicConfig(level=logging.DEBUG)

        # 确认测量规范文件存在
        spec_file = "reports/templates/PART-A001_MeasureSpec.xlsx"
        if os.path.exists(spec_file):
            print(f"✅ 测量规范文件存在: {spec_file}")
        else:
            print(f"❌ 测量规范文件不存在: {spec_file}")
            return False

        # 创建Excel导出器，使用正确的零件号
        test_filename = "test_measure_spec_final.xlsx"
        exporter = ExcelExporterEnhanced(filename=test_filename, part_no="PART-A001")

        # 创建模板
        success = exporter.create_from_template("PART-A001", "B202501", "测试员")
        if success:
            print("✅ Excel模板创建成功")
        else:
            print("❌ Excel模板创建失败")
            return False

        # 添加一些测试数据，对应不同的标准序号
        test_data = [
            (25.4, "二十五点四", "25.4"),  # 标准序号100
            (30.1, "三十点一", "30.1"),   # 标准序号200
            (28.7, "二十八点七", "28.7")  # 标准序号300
        ]

        # 设置不同的标准序号并添加数据
        for i, (val, original, processed) in enumerate(test_data):
            exporter.current_standard_id = [100, 200, 300][i]
            results = exporter.append_with_text([(val, original, processed)])
            print(f"✅ 添加数据: 标准序号={exporter.current_standard_id}, 值={val}")

        print(f"总共添加了 {len(test_data)} 条数据")

        # 执行最终格式化
        print("🔧 开始执行Excel最终格式化...")
        success = exporter.finalize_excel_file()

        if success:
            print("✅ Excel最终格式化成功")

            # 验证结果
            import openpyxl
            workbook = openpyxl.load_workbook(test_filename)
            worksheet = workbook.active

            print("\n📊 验证Excel内容:")
            found_spec_data = False
            for row in range(5, worksheet.max_row + 1):  # 从第5行开始是数据
                standard_id = worksheet.cell(row=row, column=1).value       # 标准序号
                standard_content = worksheet.cell(row=row, column=2).value   # 标准内容
                lower_limit = worksheet.cell(row=row, column=3).value      # 下限
                upper_limit = worksheet.cell(row=row, column=4).value      # 上限
                measured_value = worksheet.cell(row=row, column=6).value    # 测量值
                judgment = worksheet.cell(row=row, column=7).value         # 判断结果

                print(f"第{row}行: 标准序号={standard_id}, 测量值={measured_value}")
                print(f"  标准内容={standard_content}, 下限={lower_limit}, 上限={upper_limit}")
                print(f"  判断结果={judgment}")

                if standard_content or lower_limit or upper_limit or judgment:
                    found_spec_data = True

            workbook.close()

            if found_spec_data:
                print("\n✅ 测量规范查询功能正常工作！")
                return True
            else:
                print("\n❌ 测量规范查询功能未正常工作")
                return False
        else:
            print("❌ Excel最终格式化失败")
            return False

    except Exception as e:
        print(f"❌ 测试失败: {e}")
        import traceback
        traceback.print_exc()
        return False

    finally:
        # 清理测试文件
        if os.path.exists("test_measure_spec_final.xlsx"):
            os.remove("test_measure_spec_final.xlsx")

if __name__ == "__main__":
    success = test_measure_spec_query()
    if success:
        print("\n🎉 测量规范查询功能测试成功！")
    else:
        print("\n⚠️ 测量规范查询功能测试失败。")
    sys.exit(0 if success else 1)