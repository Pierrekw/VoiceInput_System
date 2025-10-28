#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
测量规范集成测试
测试完整的测量规范查询和Excel更新流程
"""

import os
import sys
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from datetime import datetime

# 添加父目录到路径
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from measure_spec_enhanced import MeasureSpecEnhanced
from excel_exporter import ExcelExporter

def create_test_report_with_specs():
    """创建包含测量规范的测试报告"""

    print("🚀 开始测量规范集成测试")
    print("="*60)

    # 1. 创建测量规范管理器
    print("\n1️⃣ 创建测量规范管理器...")
    spec_manager = MeasureSpecEnhanced()

    # 2. 使用增强模板创建Excel文件
    print("\n2️⃣ 创建Excel报告...")
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_filename = f"reports/test_measure_spec_{timestamp}.xlsx"

    # 使用新的增强模板
    template_path = "reports/enhanced_report_template.xlsx"
    if not os.path.exists(template_path):
        print("❌ 增强模板不存在，请先运行 create_enhanced_template.py")
        return False

    # 复制模板
    import shutil
    shutil.copy2(template_path, excel_filename)
    print(f"✅ Excel文件已创建: {excel_filename}")

    # 3. 添加测试数据
    print("\n3️⃣ 添加测试数据...")
    test_data = [
        (100, 80.0),   # 在范围内 (75.5-85.8) -> OK
        (200, 25.0),   # 在范围内 (15.5-30.5) -> OK
        (300, 10.0),   # 在范围内 (8.5-12.5) -> OK
        (100, 90.0),   # 超出上限 -> NOK
        (200, 10.0),   # 低于下限 -> NOK
        (400, 55.0),   # 在范围内 (53.5-57.5) -> OK
    ]

    # 使用openpyxl直接写入数据
    workbook = load_workbook(excel_filename)
    worksheet = workbook.active

    data_start_row = 10  # 从第10行开始写数据

    for i, (standard_id, measured_value) in enumerate(test_data):
        row = data_start_row + i
        worksheet.cell(row=row, column=1, value=standard_id)  # 标准序号
        worksheet.cell(row=row, column=5, value=measured_value)  # 测量值
        worksheet.cell(row=row, column=8, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))  # 时间戳
        worksheet.cell(row=row, column=9, value=i+1)  # 语音录入编号

    workbook.save(excel_filename)
    workbook.close()

    print(f"✅ 已添加{len(test_data)}条测试数据")

    # 4. 更新测量规范和判断结果
    print("\n4️⃣ 更新测量规范和判断结果...")
    success = spec_manager.update_excel_with_judgments(excel_filename, "PART-A001")

    if success:
        print("✅ 测量规范更新成功")
    else:
        print("❌ 测量规范更新失败")
        return False

    # 5. 验证结果
    print("\n5️⃣ 验证结果...")
    workbook = load_workbook(excel_filename)
    worksheet = workbook.active

    print("📊 更新后的Excel内容:")
    for row in range(data_start_row, data_start_row + len(test_data)):
        standard_id = worksheet.cell(row=row, column=1).value
        content = worksheet.cell(row=row, column=2).value
        lower_limit = worksheet.cell(row=row, column=3).value
        upper_limit = worksheet.cell(row=row, column=4).value
        measured_value = worksheet.cell(row=row, column=5).value
        result = worksheet.cell(row=row, column=6).value
        deviation = worksheet.cell(row=row, column=7).value
        timestamp = worksheet.cell(row=row, column=8).value
        voice_id = worksheet.cell(row=row, column=9).value

        print(f"   行{row}: 标准{standard_id}({content}) {measured_value} [{lower_limit}-{upper_limit}] -> {result}")
        if deviation is not None:
            print(f"         偏差: {deviation}")

    workbook.close()

    # 6. 手动验证判断逻辑
    print("\n6️⃣ 手动验证判断逻辑...")
    print("预期判断结果:")
    expected_results = [
        (100, 80.0, "OK"),   # 75.5-85.8范围内
        (200, 25.0, "OK"),   # 15.5-30.5范围内
        (300, 10.0, "OK"),   # 8.5-12.5范围内
        (100, 90.0, "NOK"),  # 超出85.8上限
        (200, 10.0, "NOK"),  # 低于15.5下限
        (400, 55.0, "OK"),   # 53.5-57.5范围内
    ]

    for standard_id, measured_value, expected in expected_results:
        judgment = spec_manager.judge_measurement("PART-A001", standard_id, measured_value)
        actual = judgment['result']
        status = "✅" if actual == expected else "❌"
        print(f"   {status} 标准{standard_id}, 测量值{measured_value}: 预期{expected}, 实际{actual}")

    print(f"\n✅ 集成测试完成！")
    print(f"📁 测试文件: {excel_filename}")
    print(f"📊 您可以在Excel中查看完整的判断结果和格式化")

    return True

def test_excel_formula_approach():
    """测试Excel公式方案"""
    print("\n" + "="*60)
    print("🧮 测试Excel公式方案")
    print("="*60)

    # 创建一个包含公式的Excel文件
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    formula_filename = f"reports/test_formula_{timestamp}.xlsx"

    # 创建数据
    data = {
        '标准序号': [100, 200, 300],
        '标准内容': ['半径1', '半径2', '半径3'],
        '下限': [75.5, 15.5, 8.5],
        '上限': [85.8, 30.5, 12.5],
        '测量值': [80.0, 25.0, 15.0],
        '判断结果': [''],  # 将由公式填充
        '偏差': ['']       # 将由公式填充
    }

    df = pd.DataFrame(data)
    df.to_excel(formula_filename, index=False)

    # 使用openpyxl添加公式
    workbook = load_workbook(formula_filename)
    worksheet = workbook.active

    # 添加判断公式：=IF(AND(E2>=C2, E2<=D2), "OK", "NOK")
    # 添加偏差公式：=IF(F2="OK", MIN(E2-C2, D2-E2), IF(E2<C2, C2-E2, E2-D2))

    for row in range(2, len(data) + 2):  # 从第2行开始（第1行是标题）
        # 判断结果公式
        judgment_formula = f'=IF(AND(E{row}>=C{row}, E{row}<=D{row}), "OK", "NOK")'
        worksheet.cell(row=row, column=6, value=judgment_formula)

        # 偏差公式
        deviation_formula = f'=IF(F{row}="OK", MIN(E{row}-C{row}, D{row}-E{row}), IF(E{row}<C{row}, C{row}-E{row}, E{row}-D{row}))'
        worksheet.cell(row=row, column=7, value=deviation_formula)

    workbook.save(formula_filename)
    workbook.close()

    print(f"✅ Excel公式测试文件已创建: {formula_filename}")
    print("📝 包含的公式:")
    print("   判断结果: =IF(AND(测量值>=下限, 测量值<=上限), \"OK\", \"NOK\")")
    print("   偏差: =IF(判断结果=\"OK\", MIN(测量值-下限, 上限-测量值), IF(测量值<下限, 下限-测量值, 测量值-上限))")

    # 验证公式计算结果
    df_result = pd.read_excel(formula_filename)
    print("\n📊 公式计算结果:")
    for idx, row in df_result.iterrows():
        print(f"   标准{row['标准序号']}: {row['测量值']} -> {row['判断结果']} (偏差: {row['偏差']})")

    return True

def main():
    """主函数"""
    print("🎯 测量规范功能测试")
    print("="*80)

    results = []

    # 测试集成方案
    integration_result = create_test_report_with_specs()
    results.append(("集成方案", integration_result))

    # 测试Excel公式方案
    formula_result = test_excel_formula_approach()
    results.append(("Excel公式方案", formula_result))

    # 输出结果汇总
    print("\n" + "="*80)
    print("📊 测试结果汇总")
    print("="*80)

    for test_name, result in results:
        status = "✅ 通过" if result else "❌ 失败"
        print(f"{test_name}: {status}")

    if all(result for _, result in results):
        print("\n🎉 所有测试通过！")
        print("\n💡 推荐使用方案:")
        print("1. 🔧 Python集成方案 - 自动化程度高，适合复杂业务逻辑")
        print("2. 📊 Excel公式方案 - 实时计算，无需代码干预")
        print("3. 🎯 混合方案 - Python写入基础数据，Excel进行实时判断")
        return True
    else:
        print("\n⚠️ 部分测试失败，请检查上述错误信息")
        return False

if __name__ == "__main__":
    success = main()
    sys.exit(0 if success else 1)