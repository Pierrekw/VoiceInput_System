#!/usr/bin/env python3
"""
测试Excel导出的鲁棒性 - 验证OK/NOK文本测量值的处理
"""

import sys
import os
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from excel_utils import ExcelExporterEnhanced
from openpyxl import load_workbook
import tempfile

def test_robust_measurement_handling():
    """测试鲁棒的测量值处理"""
    print("🧪 测试Excel导出的鲁棒性")
    print("=" * 50)

    try:
        # 创建临时Excel文件
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            temp_filename = tmp.name

        # 创建Excel导出器
        exporter = ExcelExporterEnhanced(
            filename=temp_filename,
            part_no="TEST-PART",
            batch_no="TEST-BATCH",
            inspector="TEST-INSPECTOR"
        )

        # 创建模板文件
        exporter.create_from_template("TEST-PART", "TEST-BATCH", "TEST-INSPECTOR")
        print("✅ Excel模板创建成功")

        # 从配置文件获取实际的special_texts配置
        from config_loader import config
        exportable_texts = config.get_exportable_texts()

        # 提取OK和NOK变体用于测试
        ok_variants = []
        nok_variants = []

        for text_config in exportable_texts:
            base_text = text_config.get('base_text', '').upper()
            variants = text_config.get('variants', [])

            if base_text == 'OK':
                ok_variants.extend(variants)
            elif base_text in ['NOK', 'NOT OK']:
                nok_variants.extend(variants)

        print(f"✅ 从配置加载OK变体: {ok_variants}")
        print(f"✅ 从配置加载NOK变体: {nok_variants}")

        # 模拟各种类型的测量数据，使用配置中的实际变体
        test_measurements = [
            (18.0, "十八", "十八"),           # 数值测量
            (31.0, "三十一", "三十一"),       # 数值测量
        ]

        # 添加配置中的OK变体
        if ok_variants:
            test_measurements.append((ok_variants[0], ok_variants[0], ok_variants[0]))

        # 添加配置中的NOK变体
        if nok_variants:
            test_measurements.append((nok_variants[0], nok_variants[0], nok_variants[0]))

        # 添加更多测试用例来验证文本标准对比
        additional_tests = []

        # 对于OK标准(300)，测试匹配和不匹配的情况
        if ok_variants:
            additional_tests.append((ok_variants[0], ok_variants[0], ok_variants[0]))  # 匹配：OK vs OK标准
        if nok_variants:
            additional_tests.append((nok_variants[0], nok_variants[0], nok_variants[0]))  # 不匹配：NOK vs OK标准

        # 对于NOK标准(500)，测试匹配和不匹配的情况
        if nok_variants:
            additional_tests.append((nok_variants[0], nok_variants[0], nok_variants[0]))  # 匹配：NOK vs NOK标准
        if ok_variants:
            additional_tests.append((ok_variants[0], ok_variants[0], ok_variants[0]))  # 不匹配：OK vs NOK标准

        # 添加其他测试用例
        additional_tests.extend([
            ("异常", "异常", "异常"),         # 异常文本
            ("unknown", "unknown", "unknown") # 未知文本
        ])

        test_measurements.extend(additional_tests)

        # 设置标准序号并添加数据
        standard_ids = [100, 200]  # 数值标准
        standard_ids.extend([300] * len(ok_variants))  # OK标准测试
        standard_ids.extend([300] * len(nok_variants))  # OK标准测试
        standard_ids.extend([500] * len(nok_variants))  # NOK标准测试
        standard_ids.extend([500] * len(ok_variants))  # NOK标准测试
        standard_ids.extend([400, 600])  # 其他测试

        for i, (value, original, processed) in enumerate(test_measurements):
            exporter.current_standard_id = standard_ids[i]
            result = exporter.append_with_text([(value, original, processed)])
            print(f"📊 添加测量 {i+1}: {value} (标准序号: {standard_ids[i]})")

        print(f"✅ 成功添加 {len(test_measurements)} 条测量数据")

        # 创建模拟的测量规范文件
        spec_filename = f"TEST-PART_MeasureSpec.xlsx"
        spec_path = os.path.join(os.path.dirname(temp_filename), "templates", spec_filename)

        # 确保templates目录存在
        os.makedirs(os.path.dirname(spec_path), exist_ok=True)

        # 创建测量规范文件
        from openpyxl import Workbook
        spec_wb = Workbook()
        spec_ws = spec_wb.active
        spec_ws.title = "测量规范"

        # 添加标题行
        spec_ws.append(["标准序号", "标准内容", "下限", "上限"])

        # 添加测试规范数据 - 包含文本标准（OK/NOK）
        spec_data = [
            (100, "尺寸测量1", 15.0, 25.0),
            (200, "尺寸测量2", 25.0, 35.0),
            (300, "OK", None, None),        # 文本标准：OK
            (400, "外观检查", None, None),  # 无数值规范
            (500, "NOK", None, None),       # 文本标准：NOK
            (600, "安全检查", None, None),  # 无数值规范
        ]

        for row in spec_data:
            spec_ws.append(row)

        spec_wb.save(spec_path)
        spec_wb.close()
        print(f"✅ 创建测量规范文件: {spec_path}")

        # 应用最终格式化
        success = exporter.finalize_excel_file()
        if success:
            print("✅ Excel最终格式化成功")
        else:
            print("❌ Excel最终格式化失败")
            return False

        # 验证Excel文件内容
        print("\n🔍 验证Excel文件内容:")
        wb = load_workbook(temp_filename)
        ws = wb.active

        # 检查数据行
        data_start_row = 5  # 从第5行开始是数据
        for i in range(len(test_measurements)):
            row = data_start_row + i
            standard_id = ws.cell(row=row, column=1).value
            content = ws.cell(row=row, column=2).value
            lower_limit = ws.cell(row=row, column=3).value
            upper_limit = ws.cell(row=row, column=4).value
            measured_value = ws.cell(row=row, column=6).value
            judgment = ws.cell(row=row, column=7).value
            deviation = ws.cell(row=row, column=8).value

            print(f"\n📋 行 {row}:")
            print(f"  标准序号: {standard_id}")
            print(f"  标准内容: {content}")
            print(f"  上下限: {lower_limit} - {upper_limit}")
            print(f"  测量值: {measured_value}")
            print(f"  判断结果: {judgment}")
            print(f"  偏差: {deviation}")

            # 验证关键信息是否填写
            checks = [
                (standard_id is not None, "标准序号"),
                (content is not None, "标准内容"),
                (judgment is not None, "判断结果"),
                (deviation is not None, "偏差")
            ]

            for check, desc in checks:
                if check:
                    print(f"    ✅ {desc}: 已填写")
                else:
                    print(f"    ❌ {desc}: 缺失")

        wb.close()

        # 清理临时文件
        os.unlink(temp_filename)
        if os.path.exists(spec_path):
            os.unlink(spec_path)

        print("\n🎉 鲁棒性测试完成!")
        return True

    except Exception as e:
        print(f"❌ 测试失败: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    print("🚀 开始Excel导出鲁棒性测试")
    print("=" * 60)
    test_robust_measurement_handling()