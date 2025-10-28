#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
检查最终生成的报告文件
"""

import pandas as pd
from openpyxl import load_workbook

def check_final_report():
    """检查最终生成的报告文件"""
    print("🔍 检查最终生成的报告文件")
    print("="*60)

    # 重新生成一个测试文件
    from excel_exporter_enhanced import ExcelExporterEnhanced

    # 创建测量规范文件
    spec_data = [
        [100, '半径1', 75.5, 85.8],
        [200, '半径2', 15.5, 30.5],
    ]

    spec_df = pd.DataFrame(spec_data, columns=['标准序号', '标准内容', '下限', '上限'])
    spec_filename = "reports/FINAL-CHECK_MeasureSpec.xlsx"

    import os
    os.makedirs("reports", exist_ok=True)
    spec_df.to_excel(spec_filename, index=False)

    # 创建导出器
    exporter = ExcelExporterEnhanced(
        filename="reports/final_check.xlsx",
        part_no="FINAL-CHECK",
        batch_no="B001",
        inspector="最终检查员"
    )

    # 使用模板创建并写入数据
    exporter.create_from_template("FINAL-CHECK", "B001", "最终检查员")

    test_data = [(100, 80.0, "半径1 八十")]
    exporter.append_with_text(test_data)

    exporter.finalize_excel_file()

    # 检查生成的文件
    file_path = "reports/final_check.xlsx"
    if os.path.exists(file_path):
        print(f"📊 检查文件: {file_path}")

        workbook = load_workbook(file_path)
        worksheet = workbook.active

        print(f"\n📋 完整文件内容:")
        for row in range(1, min(8, worksheet.max_row + 1)):
            row_content = []
            for col in range(1, min(11, worksheet.max_column + 1)):
                cell_value = worksheet.cell(row=row, column=col).value
                if cell_value is not None:
                    cell_value = str(cell_value)[:40]
                else:
                    cell_value = "None"
                row_content.append(f"{cell_value:>15}")
            print(f"行{row:2d}: {' | '.join(row_content)}")

        workbook.close()
        os.remove(file_path)
        print(f"\n🧹 已清理测试文件")

    # 清理测量规范文件
    if os.path.exists(spec_filename):
        os.remove(spec_filename)

if __name__ == "__main__":
    check_final_report()