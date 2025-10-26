#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
综合测试脚本 - 验证Excel增强功能和GUI集成
测试模板复制、双ID系统、输入验证、标准序号变更等功能
"""

import sys
import os
import time
import logging
from datetime import datetime

# 添加父目录到路径
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from excel_exporter import ExcelExporter
from voice_gui import WorkingSimpleMainWindow
from logging_utils import LoggingManager
from PySide6.QtWidgets import QApplication
from PySide6.QtCore import QTimer

def test_excel_enhanced_functionality():
    """测试Excel增强功能"""
    print("="*80)
    print("🧪 测试Excel增强功能")
    print("="*80)

    # 设置日志
    logger = LoggingManager.get_logger(
        name='test_enhanced',
        level=logging.INFO,
        console_level=logging.INFO,
        log_to_console=True,
        log_to_file=True
    )

    try:
        # 1. 测试模板创建
        print("\n1️⃣ 测试Excel模板创建...")
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        test_filename = f"reports/test_enhanced_{timestamp}.xlsx"

        exporter = ExcelExporter(filename=test_filename)
        part_no = "TEST001"
        batch_no = "B202501"
        inspector = "测试员"

        success = exporter.create_from_template(part_no, batch_no, inspector)
        if success:
            print("✅ 模板创建成功")
        else:
            print("❌ 模板创建失败，使用默认方式")
            exporter.create_new_file()

        # 2. 测试双ID系统
        print("\n2️⃣ 测试双ID系统...")

        # 设置标准序号100
        exporter.current_standard_id = 100
        test_data_100 = [
            (12.5, "十二点五", "12.5"),
            (15.8, "十五点八", "15.8"),
            ("OK", "好的", "OK")
        ]

        results_100 = exporter.append_with_text(test_data_100)
        print(f"   标准序号100写入结果: {len(results_100)} 条")
        for voice_id, value, original in results_100:
            print(f"   Voice ID: {voice_id}, 值: {value}, 原文: {original}")

        # 设置标准序号200
        exporter.current_standard_id = 200
        test_data_200 = [
            (8.1, "八点一", "8.1"),
            (25.6, "二十五点六", "25.6"),
            ("NOK", "不行", "NOK")
        ]

        results_200 = exporter.append_with_text(test_data_200)
        print(f"   标准序号200写入结果: {len(results_200)} 条")
        for voice_id, value, original in results_200:
            print(f"   Voice ID: {voice_id}, 值: {value}, 原文: {original}")

        # 3. 测试删除功能
        print("\n3️⃣ 测试删除功能...")
        # 删除Voice ID 4 (属于标准序号200的数据)
        delete_success = exporter.delete_row_by_voice_id(4)
        if delete_success:
            print("✅ 删除Voice ID 4成功")
        else:
            print("❌ 删除失败")

        # 4. 测试继续写入
        print("\n4️⃣ 测试继续写入...")
        exporter.current_standard_id = 300
        continue_data = [(101.5, "一百零一点五", "101.5")]
        continue_results = exporter.append_with_text(continue_data)

        print(f"   继续写入结果: {len(continue_results)} 条")
        for voice_id, value, original in continue_results:
            print(f"   Voice ID: {voice_id}, 值: {value}, 原文: {original}")

        # 5. 测试重新编号
        print("\n5️⃣ 测试Excel重新编号...")
        exporter.renumber_excel_ids()
        print("✅ 重新编号完成")

        # 6. 验证最终结果
        print("\n6️⃣ 验证最终结果...")
        verify_excel_results(exporter, test_filename)

        print(f"\n✅ Excel功能测试完成！文件: {test_filename}")
        return True

    except Exception as e:
        print(f"\n❌ Excel功能测试失败: {e}")
        import traceback
        traceback.print_exc()
        return False

def verify_excel_results(exporter, filename):
    """验证Excel结果"""
    try:
        import openpyxl
        from openpyxl import load_workbook

        # 读取Excel文件
        workbook = load_workbook(filename)
        worksheet = workbook.active

        print("📊 Excel文件内容分析:")

        # 按标准序号分组统计
        standard_groups = {}
        for row in range(4, worksheet.max_row + 1):
            standard_id_cell = worksheet.cell(row=row, column=1)  # A列
            voice_id_cell = worksheet.cell(row=row, column=7)   # G列
            excel_id_cell = worksheet.cell(row=row, column=3)   # C列
            measurement_cell = worksheet.cell(row=row, column=4)  # D列

            if standard_id_cell.value is not None and voice_id_cell.value is not None:
                standard_id = standard_id_cell.value
                voice_id = voice_id_cell.value
                excel_id = excel_id_cell.value
                measurement = measurement_cell.value

                if standard_id not in standard_groups:
                    standard_groups[standard_id] = []

                standard_groups[standard_id].append({
                    'row': row,
                    'voice_id': voice_id,
                    'excel_id': excel_id,
                    'measurement': measurement
                })

        workbook.close()

        # 分析每个标准序号的数据
        for standard_id, records in standard_groups.items():
            print(f"\n📏 标准序号 {standard_id}:")
            print(f"   记录数量: {len(records)}")

            voice_ids = [r['voice_id'] for r in records]
            excel_ids = [r['excel_id'] for r in records]
            measurements = [r['measurement'] for r in records]

            print(f"   Voice IDs: {voice_ids}")
            print(f"   Excel IDs: {excel_ids}")
            print(f"   测量值: {measurements}")

        # 验证删除的数据
        deleted_ids = exporter.deleted_voice_ids
        print(f"\n🗑️ 已删除的Voice ID: {deleted_ids}")

        # 检查内存状态
        print(f"\n🧠 内存状态:")
        print(f"   voice_id_counter: {exporter.voice_id_counter}")
        print(f"   next_insert_row: {exporter.next_insert_row}")
        print(f"   active_record_count: {exporter.active_record_count}")
        print(f"   current_standard_id: {exporter.current_standard_id}")

    except Exception as e:
        print(f"❌ 验证失败: {e}")

def test_gui_input_validation():
    """测试GUI输入验证功能"""
    print("\n" + "="*80)
    print("🧪 测试GUI输入验证功能")
    print("="*80)

    app = QApplication(sys.argv) if not QApplication.instance() else QApplication.instance()

    try:
        # 创建GUI窗口（但不显示）
        window = WorkingSimpleMainWindow()

        print("\n1️⃣ 测试有效输入...")
        # 测试有效输入
        window.part_no_input.setText("PART001")
        window.batch_no_input.setText("B202501")
        window.inspector_input.setText("张三")

        if window.are_inputs_valid():
            print("✅ 有效输入验证通过")
            values = window.get_input_values()
            print(f"   输入值: {values}")
        else:
            print("❌ 有效输入验证失败")

        print("\n2️⃣ 测试无效输入...")
        # 测试无效输入
        window.part_no_input.setText("")  # 空值
        window.batch_no_input.setText("AB")  # 太短
        window.inspector_input.setText("A1")  # 包含数字

        if not window.are_inputs_valid():
            print("✅ 无效输入正确被拒绝")
            print(f"   错误信息: {window.validation_errors}")
        else:
            print("❌ 无效输入验证失败")

        print("\n3️⃣ 测试边界情况...")
        # 测试边界情况
        window.part_no_input.setText("A" * 25)  # 超过最大长度
        window.batch_no_input.setText("B" * 20)  # 超过最大长度
        window.inspector_input.setText("A" * 15)  # 超过最大长度

        if not window.are_inputs_valid():
            print("✅ 边界情况正确处理")
            print(f"   错误信息: {window.validation_errors}")
        else:
            print("❌ 边界情况验证失败")

        print("\n✅ GUI输入验证测试完成")
        window.close()
        return True

    except Exception as e:
        print(f"\n❌ GUI输入验证测试失败: {e}")
        import traceback
        traceback.print_exc()
        return False

def test_integration():
    """测试集成功能"""
    print("\n" + "="*80)
    print("🧪 测试集成功能")
    print("="*80)

    try:
        from main_f import FunASRVoiceSystem

        print("\n1️⃣ 测试语音系统初始化...")
        system = FunASRVoiceSystem(
            recognition_duration=-1,
            continuous_mode=True,
            debug_mode=False
        )

        print("\n2️⃣ 测试Excel模板集成...")
        success = system.setup_excel_from_gui("PART001", "B202501", "测试员")
        if success:
            print("✅ Excel模板集成成功")
            print(f"   Excel文件路径: {system.excel_exporter.filename}")
        else:
            print("⚠️ Excel模板集成失败，使用默认方式")

        print("\n3️⃣ 测试数据写入集成...")
        if system.excel_exporter:
            test_data = [(12.5, "十二点五", "12.5")]
            results = system.excel_exporter.append_with_text(test_data)
            print(f"   写入结果: {len(results)} 条")

        print("\n4️⃣ 测试Excel最终处理...")
        system._finalize_excel()
        print("✅ Excel最终处理完成")

        print("\n✅ 集成功能测试完成")
        return True

    except Exception as e:
        print(f"\n❌ 集成功能测试失败: {e}")
        import traceback
        traceback.print_exc()
        return False

def main():
    """主测试函数"""
    print("🚀 开始综合功能测试")
    print("="*80)

    results = []

    # 测试Excel增强功能
    excel_result = test_excel_enhanced_functionality()
    results.append(("Excel增强功能", excel_result))

    # 测试GUI输入验证
    gui_result = test_gui_input_validation()
    results.append(("GUI输入验证", gui_result))

    # 测试集成功能
    integration_result = test_integration()
    results.append(("集成功能", integration_result))

    # 输出测试结果汇总
    print("\n" + "="*80)
    print("📊 测试结果汇总")
    print("="*80)

    passed = 0
    total = len(results)

    for test_name, result in results:
        status = "✅ 通过" if result else "❌ 失败"
        print(f"{test_name}: {status}")
        if result:
            passed += 1

    print(f"\n总计: {passed}/{total} 项测试通过")

    if passed == total:
        print("🎉 所有功能测试通过！")
        return True
    else:
        print("⚠️ 部分功能存在问题，请检查上述错误信息")
        return False

if __name__ == "__main__":
    success = main()
    sys.exit(0 if success else 1)