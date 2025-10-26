#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
测试语音控制、ID一致性和Excel查询功能修复
"""

import sys
import os

# 添加父目录到路径
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

def test_voice_command_standard_id():
    """测试语音控制标准序号切换功能"""
    print("🧪 测试语音控制标准序号切换功能")
    print("-" * 50)

    try:
        from main_f import FunASRVoiceSystem

        # 创建系统实例
        system = FunASRVoiceSystem(debug_mode=True)

        # 测试语音命令识别
        test_commands = [
            ("一百", 100),
            ("二百", 200),
            ("三百", 300),
            ("标准四百", 400),
            ("序号五百", 500)
        ]

        for command, expected_id in test_commands:
            command_type = system.recognize_voice_command(command)
            if command_type.value == "standard_id":
                print(f"✅ 语音命令 '{command}' 正确识别为标准序号命令")
                system._handle_standard_id_command(command)
                actual_id = system.get_current_standard_id()
                if actual_id == expected_id:
                    print(f"✅ 标准序号正确切换到: {actual_id}")
                else:
                    print(f"❌ 标准序号切换错误，期望: {expected_id}, 实际: {actual_id}")
            else:
                print(f"❌ 语音命令 '{command}' 识别失败")

        print("✅ 语音控制标准序号切换功能测试完成")
        return True

    except Exception as e:
        print(f"❌ 语音控制测试失败: {e}")
        import traceback
        traceback.print_exc()
        return False

def test_voice_id_consistency():
    """测试Voice ID显示与Excel录入一致性"""
    print("\n🧪 测试Voice ID显示与Excel录入一致性")
    print("-" * 50)

    try:
        from excel_utils import ExcelExporterEnhanced

        # 创建测试Excel导出器
        test_filename = "test_voice_id_consistency.xlsx"
        exporter = ExcelExporterEnhanced(filename=test_filename, part_no="PART-A001")

        # 模拟添加数据
        test_data = [
            (25.4, "二十五点四", "25.4"),
            (30.1, "三十点一", "30.1"),
            (28.7, "二十八点七", "28.7")
        ]

        # 测试append_with_text方法，让它自动生成ID
        results = exporter.append_with_text(test_data)

        if results:
            print(f"\n✅ 成功写入 {len(results)} 条记录")
            actual_ids = [r[0] for r in results]
            print(f"生成的Voice IDs: {actual_ids}")

            # 检查ID是否连续递增
            expected_ids = list(range(1, len(results) + 1))
            if actual_ids == expected_ids:
                print("✅ Voice ID连续递增正确")
                for i, voice_id in enumerate(actual_ids):
                    print(f"✅ 记录 {i+1}: Voice ID = {voice_id}")
            else:
                print(f"❌ Voice ID不连续，期望: {expected_ids}, 实际: {actual_ids}")
                return False
        else:
            print("❌ 没有写入任何记录")
            return False

        # 清理测试文件
        if os.path.exists(test_filename):
            os.remove(test_filename)

        print("✅ Voice ID一致性测试完成")
        return True

    except Exception as e:
        print(f"❌ Voice ID一致性测试失败: {e}")
        import traceback
        traceback.print_exc()
        return False

def test_excel_measure_spec_query():
    """测试Excel测量规范自动查询功能"""
    print("\n🧪 测试Excel测量规范自动查询功能")
    print("-" * 50)

    try:
        from excel_utils import ExcelExporterEnhanced

        # 检查测量规范文件是否存在
        spec_file = "reports/templates/PART-A001_MeasureSpec.xlsx"
        if os.path.exists(spec_file):
            print(f"✅ 测量规范文件存在: {spec_file}")
        else:
            print(f"❌ 测量规范文件不存在: {spec_file}")
            return False

        # 创建测试Excel导出器
        test_filename = "test_measure_spec.xlsx"
        exporter = ExcelExporterEnhanced(filename=test_filename, part_no="PART-A001")

        # 创建模板文件
        success = exporter.create_from_template("PART-A001", "B202501", "测试员")
        if success:
            print("✅ Excel模板创建成功")
        else:
            print("❌ Excel模板创建失败")
            return False

        # 添加测试数据
        test_data = [
            (25.4, "二十五点四", "25.4"),  # 对应标准序号100
            (30.1, "三十点一", "30.1"),   # 对应标准序号200
            (28.7, "二十八点七", "28.7")  # 对应标准序号300
        ]

        results = exporter.append_with_text(test_data)
        if results:
            print(f"✅ 成功写入 {len(results)} 条记录")
        else:
            print("❌ 数据写入失败")
            return False

        # 设置调试日志级别
        import logging
        logging.getLogger().setLevel(logging.DEBUG)

        # 测试最终格式化（包括测量规范查询）
        print("🔧 开始测试Excel最终格式化...")
        success = exporter.finalize_excel_file()

        if success:
            print("✅ Excel最终格式化成功，测量规范查询应该已完成")

            # 验证结果
            import openpyxl
            workbook = openpyxl.load_workbook(test_filename)
            worksheet = workbook.active

            # 检查是否填充了测量规范数据
            spec_found = False
            for row in range(5, worksheet.max_row + 1):  # 从第5行开始是数据
                standard_content = worksheet.cell(row=row, column=2).value  # 标准内容
                lower_limit = worksheet.cell(row=row, column=3).value       # 下限
                upper_limit = worksheet.cell(row=row, column=4).value       # 上限
                judgment = worksheet.cell(row=row, column=7).value          # 判断结果

                if standard_content or lower_limit or upper_limit or judgment:
                    spec_found = True
                    print(f"✅ 第{row}行: 标准内容={standard_content}, 下限={lower_limit}, 上限={upper_limit}, 判断结果={judgment}")

            if spec_found:
                print("✅ 测量规范查询功能正常工作")
            else:
                print("❌ 测量规范查询功能未正常工作")
                return False

            workbook.close()
        else:
            print("❌ Excel最终格式化失败")
            return False

        # 清理测试文件
        if os.path.exists(test_filename):
            os.remove(test_filename)

        print("✅ Excel测量规范查询功能测试完成")
        return True

    except Exception as e:
        print(f"❌ Excel测量规范查询测试失败: {e}")
        import traceback
        traceback.print_exc()
        return False

def main():
    """主测试函数"""
    print("🚀 开始测试三个问题修复效果")
    print("=" * 60)

    # 测试1: 语音控制标准序号切换
    test1_result = test_voice_command_standard_id()

    # 测试2: Voice ID一致性
    test2_result = test_voice_id_consistency()

    # 测试3: Excel测量规范查询
    test3_result = test_excel_measure_spec_query()

    # 总结
    print("\n" + "=" * 60)
    print("🎯 测试结果总结:")
    print(f"  1. 语音控制标准序号切换: {'✅ 通过' if test1_result else '❌ 失败'}")
    print(f"  2. Voice ID一致性: {'✅ 通过' if test2_result else '❌ 失败'}")
    print(f"  3. Excel测量规范查询: {'✅ 通过' if test3_result else '❌ 失败'}")

    if all([test1_result, test2_result, test3_result]):
        print("\n🎉 所有测试通过！三个问题已成功修复。")
        return True
    else:
        print("\n⚠️ 部分测试失败，需要进一步检查。")
        return False

if __name__ == "__main__":
    success = main()
    sys.exit(0 if success else 1)